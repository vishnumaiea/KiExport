#=============================================================================================#

# KiExport
# Tool to export manufacturing files from KiCad PCB projects.
# Author: Vishnu Mohanan (@vishnumaiea, @vizmohanan)
# Version: 0.1.9
# Last Modified: +05:30 13:15:20 PM 05-10-2025, Sunday
# GitHub: https://github.com/vishnumaiea/KiExport
# License: MIT
#
# Update 1.9 22/08/2025 by Leor Weinstein
# 1. Modified default config json
#      Added additional CLI commands as available on Kicad CLI 9.0.3 to be supported by JSON configuration file
#      Added support for multylayer boards In1.cu-In8.cu
# 2. Corrected support for pcb_renders preset string from CLI
#      replaced \w in bare word pattern search with [\w-] so will allow hyphen as part of word string
# 3. Removed the inclusion of gbrjob in the created gerber zip file since most mfc factories will not use it 
#      (it was also not functional anyhow see below correction)
# 4. Added correction to gerber file name in gbrjob to match the updated single gerber file names 
#      since these files were renamed with ver string and can no longer be accessed by gbrjob directly
# 5. When merging all pcb pdf file to a single file, this is now done in the order listed under layer list
#      that makes more sense than just a random OS files order.
# 6. Added delete of all single page pdf files created from pcb_pdf and only keeped the final merged pdf
#      As such removed also the zip creation of the individual pdf files since only the merged pdf is needed
# 7. Also forced drill marking to None (0) on Silk and Fab layers when generating pcb_pdf
# 8. Added new cmd "pcb_fab_pdf", this is similar to pcb_pdf only specific behaviour for F.fab and B.fab layers only
#      it will only accept either F.fab and B.fab in layer list and ignore all others
#      if F.fab is present it will automatically generate it with mirror false
#      if B.fab is present it will automatically generate it with mirror true
#      if both F.fab and B.fab were generated it will create a single combined fab pdf under the PCB-Fab dir
# 9. Added special handling for cmd "pcb_pdf", this will have the following specific behaviour
#      All Front layers (Cu,Silkscreen,Mask,Paste,Fab,Courtyard) if present will automatically generate it with mirror false
#      All Bottom layers (Cu,Silkscreen,Mask,Paste,Fab,Courtyard) if present will automatically generate it with mirror true
#      so can easily compare to PCB view looking on it from bottom side
#10. Added -c or --clean command line parameter
#      This will remove completely the Kicad project revision directory so will generate final clean output
#      of all exports, thus removing any seq accumulated dates and files along the way.
#      This is useful for producing one final approved release of all files for current revision without the confusion of older runs for same version
#11. At first I thought PCB and Schematic may have different revisions 
#      (e.g. parts value change or patches not corrected in PCB design) and it can get confusing which rev dir 
#      is for which, so I thought to create an additional hierarchy: PCB and SCH.
#      But then I realized that this is not the case, since any change of schematic will affect BOM and fab layers in PCB
#      So better to have a unique global version identical to PCB and schematic (project version).
#      However, It is possible that gerber files for remaking PCB production WILL NOT be made for each version
#      (e.g. in case of value change and wire patches), so gereber files will only be found in specific versions dirs
#      for which PCB production was generated.
#      It is therfore recommended that user wll keep a unique single version number for sch and pcb files.
#      Warning will be provided if schematic and PCB have different revisions
#12. Bug in create_final_directory was reversly selecting using config or cli directory
#13. In case of requested in cmd config to generate both XLS and CSV, modified code so will only generate one CSV for both
#14. Added sch_erc command option with dedictaed parameter structure similar to pcb_drc
#15. Added additional argument to XLS gnerating BOM "kie_row_height" this will allow user to easily set the desired value for row height in generated XLS.
#      Value of 0 or None or none digit will set row to auto height, otherwise will be set to value specified 
#16. The custom_widths dictionary for bom xls generation was changed to be loaded from an the configuration file.
#      Added additional dictionary object to XLS gnerating BOM "kie_custom_width:{...}"
#      If specified will use the key:value pairs in it for XLS width defintion
#      If no valid found and loaded than using the predefined custom_width in python script
#17. Added additional dictionary object to XLS gnerating BOM "custom_styles:{...}"
#      If specified will allow setting specific stypes to specific columns (most notably Hyperlink style to URL columns
#      If no valid found and loaded than using the predefined custom_styles in python script (Link,Datasheet and URL as Hyperlink)
#18. PCB-render default set of renders was completely changed, not sure why had PCB-Part, PCB-Pads, PCB-Paste?
#      Also was placing Part/Pads/Paste value in Preset flag although this flag does not accept such values ??!!
#      Created simple defaults: Top,Bottom,Left,Right,Front,Back and Perspective for a perspective view of the board.
#      Kicad 9.0.3 still has a bug in handling preset for 3D renderer and output all layers even unneeded ones (like user drawing)
#      perhaps when fixed in V10 or sooner can select preset layer name here
#19. If project name not specified in configuration file or is set to "" or "None" or "Auto" value
#      will search for kicad_pcb and kicad_sch files if found one of each with identical names will use this as project name
#      This will allow easy re-use of same config across multiple projects since usually only one sch/pcb project found in same dir
#20. Added source_zip command in run command list options
#      This will be create a SRC directory in specified output directory and will zip to it all relevant source file, 
#      This can be useful for backup of actual sources used to produce a specific final release. 
#      Source file included are by default: kicad_pcb, kicad_sch, kicad_pro, kicad_prl, net.
#      Above project name extensions to inlcude in zip may be defined unser the "source_zip" command in config file
#      This will only be executed at the end of all generations and only if successful generation of all commands in run 
#      This is usually to be used together with --clean option for final clean release generation
#      Zip subroutines were slightly modified for above usage
#21. Note! Known issue with "kie_single_file" under pcb_pdf, the code seem to be remarked and assumed this 
#      flag is always set, not modified since in any case this is the preferred operation.   
#=============================================================================================#

import subprocess
import argparse
import os
import shutil
import stat
import re
from datetime import datetime
import zipfile
import json
import pymupdf
import ast
import sys
import semver
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.styles import PatternFill

#=============================================================================================#

APP_NAME = "KiExport"
APP_VERSION = "0.1.9"
APP_DESCRIPTION = "Tool to export manufacturing files from KiCad PCB projects."
APP_AUTHOR = "Vishnu Mohanan (@vishnumaiea, @vizmohanan)"

SAMPLE_PCB_FILE = "Mitayi-Pico-D1/Mitayi-Pico-RP2040.kicad_pcb"
MIN_CONFIG_JSON_VERSION = "1.6"  # Minimum required version of the config JSON file
MIN_KICAD_VERSION = "8.0"  # Minimum required version of the config JSON file

current_config = None
default_config = None

logger = None  # Logger instance
logPath = None  # Path to the log file

command_exec_status = {}  # Command execution status

DEFAULT_CONFIG_JSON = '''
{
  "name": "KiExport.JSON",
  "description": "Configuration file for KiExport",
  "filetype": "json",
  "version": "1.9",
  "project_name": "Project Name",
  "commands": [
    "pcb_drc",
    "gerbers",
    "sch_erc",
    "sch_pdf",
    ["bom", "CSV"],
    ["bom", "XLS"],
    ["bom", "HTML"],
    "pcb_pdf",
    "pcb_fab_pdf",
    "positions",
    "svg",
    ["pcb_render", "PCB-Perspective"],
    ["pcb_render", "PCB-Top"],
    ["pcb_render", "PCB-Bottom"],
    ["pcb_render", "PCB-Front"],
    ["pcb_render", "PCB-Back"],
    ["pcb_render", "PCB-Left"],
    ["pcb_render", "PCB-Right"],
    ["ddd", "STEP"],
    ["ddd", "VRML"],
    "source_zip"
  ],
  "kicad_cli_path": "C:\\\\Program Files\\\\KiCad\\\\9.99\\bin\\\\kicad-cli.exe",
  "kicad_python_path": "C:\\\\Program Files\\\\KiCad\\\\9.99\\\\bin\\\\python.exe",
  "ibom_path": "%USERPROFILE%\\\\Documents\\\\KiCad\\\\9.99\\\\3rdparty\\\\plugins\\\\org_openscopeproject_InteractiveHtmlBom\\\\generate_interactive_bom.py",
  "kiexport_log_path": "Export\\\\kiexport.log",
  "data": {
    "pcb_drc": {
      "--output_dir": "Export",
      "--define-var": false,
      "--format": "report",
      "--all-track-errors": false,
      "--schematic-parity": true,
      "--units": "mm",
      "--severity-all": false,
      "--severity-error": true,
      "--severity-warning": true,
      "--severity-exclusions": false,
      "--exit-code-violations": false
    },
    "gerbers": {
      "--output_dir": "Export",
      "--layers": [
        "F.Cu",
        "In1.Cu",
        "In2.Cu",
        "In3.Cu",
        "In4.Cu",
        "In5.Cu",
        "In6.Cu",
        "In7.Cu",
        "In8.Cu",
        "B.Cu",
        "F.Paste",
        "B.Paste",
        "F.Silkscreen",
        "B.Silkscreen",
        "F.Mask",
        "B.Mask",
        "User.Drawings",
        "User.Comments",
        "Edge.Cuts",
        "F.Courtyard",
        "B.Courtyard",
        "F.Fab",
        "B.Fab"
      ],
      "--drawing-sheet": false,
      "--define-var": false,
      "--exclude-refdes": false,
      "--exclude-value": false,
      "--include-border-title": false,
      "--sketch-pads-on-fab-layers": false,
      "--hide-DNP-footprints-on-fab-layers": false,
      "--sketch-DNP-footprints-on-fab-layers": false,
      "--crossout-DNP-footprints-on-fab-layers": false,
      "--no-x2": false,
      "--no-netlist": true,
      "--subtract-soldermask": false,
      "--disable-aperture-macros": false,
      "--use-drill-file-origin": true,
      "--precision": 6,
      "--no-protel-ext": true,
      "--plot-invisible-text": false,
      "--common-layers": false,
      "--board-plot-params": false,
      "kie_include_drill": false
    },
    "drills": {
      "--output_dir": "Export",
      "--format": "excellon",
      "--drill-origin": "plot",
      "--excellon-zeros-format": "decimal",
      "--excellon-oval-format": "route",
      "--excellon-units": "mm",
      "--excellon-mirror-y": false,
      "--excellon-min-header": false,
      "--excellon-separate-th": false,
      "--generate-map": true,
      "--map-format": "pdf",
      "--gerber-precision": false
    },
    "bom": {
      "CSV": {
        "--output_dir": "Export",
        "--preset": "Group by MPN-DNP",
        "--format-preset": "CSV",
        "--fields": "${ITEM_NUMBER},Reference,Value,Name,Footprint,${QUANTITY},${DNP},MPN,MFR,Alt MPN",
        "--labels": "#,Reference,Value,Name,Footprint,Qty,DNP,MPN,MFR,Alt MPN",
        "--group-by": "${DNP},MPN",
        "--sort-field": "Reference",
        "--sort-asc": true,
        "--filter": false,
        "--exclude-dnp": false,
        "--include-excluded-from-bom": false,
        "--field-delimiter": false,
        "--string-delimiter": false,
        "--ref-delimiter": false,
        "--ref-range-delimiter": false,
        "--keep-tabs": false,
        "--keep-line-breaks": false
      },
      "XLS": {
        "--output_dir": "Export",
        "kie_row_height": 0,
        "kie_custom_widths": {
            "#": 6,
            "Reference": 34,
            "Value": 34,
            "Qty": 6,
            "DNP": 6,
            "Description": 40,
            "MFR": 24,
            "MPN": 32,
            "Alt MPN": 32
        },
        "kie_custom_styles": {
            "Link": "Hyperlink",
            "Datasheet": "Hyperlink",
            "URL": "Hyperlink"
        }
      },
      "HTML": {
        "--output_dir": "Export",
        "--show-dialog": false,
        "--dark-mode": true,
        "--hide-pads": false,
        "--show-fabrication": false,
        "--hide-silkscreen": false,
        "--highlight-pin1": "all",
        "--no-redraw-on-drag": false,
        "--board-rotation": "0",
        "--offset-back-rotation": "0",
        "--checkboxes": "Source,Placed",
        "--bom-view": "left-right",
        "--layer-view": "FB",
        "--no-compression": false,
        "--no-browser": true,
        "--include-tracks": true,
        "--include-nets": true,
        "--sort-order": "C,R,L,D,U,Y,X,F,SW,A,~,HS,CNN,J,P,NT,MH",
        "--blacklist": false,
        "--no-blacklist-virtual": false,
        "--blacklist-empty-value": false,
        "--netlist-file": false,
        "--extra-data-file": true,
        "--extra-fields": false,
        "--show-fields": "Value,Footprint,Name,MPN,MFR,Alt MPN",
        "--group-fields": "Value,Footprint,MPN",
        "--normalize-field-case": false,
        "--variant-fields": false,
        "--variants-whitelist": false,
        "--dnp-field": false
      }
    },
    "sch_erc": {
      "--output_dir": "Export",
      "--define-var": false,
      "--format": "report",
      "--units": "mm",
      "--severity-all": false,
      "--severity-error": true,
      "--severity-warning": true,
      "--severity-exclusions": false,
      "--exit-code-violations": false
    },
    "sch_pdf": {
      "--output_dir": "Export",
      "--drawing-sheet": false,
      "--define-var": false,
      "--theme": "User",
      "--black-and-white": false,
      "--exclude-drawing-sheet": false,
      "--default-font": false,
      "--exclude-pdf-property-popups": false,
      "--exclude-pdf-hierarchical-links": false,
      "--exclude-pdf-metadata": false,
      "--no-background-color": false,
      "--pages": false
    },
    "pcb_pdf": {
      "--output_dir": "Export",
      "--layers": [
        "F.Cu",
        "In1.Cu",
        "In2.Cu",
        "In3.Cu",
        "In4.Cu",
        "In5.Cu",
        "In6.Cu",
        "In7.Cu",
        "In8.Cu",
        "B.Cu",
        "F.Paste",
        "B.Paste",
        "F.Silkscreen",
        "B.Silkscreen",
        "F.Mask",
        "B.Mask",
        "User.Drawings",
        "User.Comments",
        "Edge.Cuts",
        "F.Courtyard",
        "B.Courtyard",
        "F.Fab",
        "B.Fab"
      ],
      "kie_common_layers": ["Edge.Cuts"],
      "--drawing-sheet": false,
      "--define-var": false,
      "--mirror": false,
      "--exclude-refdes": false,
      "--exclude-value": false,
      "--include-border-title": true,
      "--subtract-soldermask": false,
      "--sketch-pads-on-fab-layers": false,
      "--hide-DNP-footprints-on-fab-layers": false,
      "--sketch-DNP-footprints-on-fab-layers": false,
      "--crossout-DNP-footprints-on-fab-layers": false,
      "--negative": false,
      "--black-and-white": false,
      "--theme": "User",
      "--drill-shape-opt": 2,
      "--plot-invisible-text": false,
      "--mode-single": false,
      "--mode-separate": false,
      "--mode-multipage": false,
      "kie_single_file": true
    },
    "pcb_fab_pdf": {
      "--output_dir": "Export",
      "--layers": [
        "F.Fab",
        "B.Fab"
      ],
      "kie_common_layers": ["Edge.Cuts"],
      "--drawing-sheet": false,
      "--define-var": false,
      "--mirror": false,
      "--exclude-refdes": false,
      "--exclude-value": false,
      "--include-border-title": true,
      "--subtract-soldermask": false,
      "--sketch-pads-on-fab-layers": false,
      "--hide-DNP-footprints-on-fab-layers": false,
      "--sketch-DNP-footprints-on-fab-layers": false,
      "--crossout-DNP-footprints-on-fab-layers": false,
      "--negative": false,
      "--black-and-white": false,
      "--theme": "User",
      "--drill-shape-opt": 2,
      "--plot-invisible-text": false,
      "--mode-single": false,
      "--mode-separate": false,
      "--mode-multipage": false,
      "kie_single_file": true
    },
    "positions": {
      "--output_dir": "Export",
      "--side": "front,back,both",
      "--format": "csv",
      "--units": "mm",
      "--bottom-negate-x": false,
      "--use-drill-file-origin": true,
      "--smd-only": false,
      "--exclude-fp-th": false,
      "--exclude-dnp": false,
      "--gerber-board-edge": false
    },
    "svg": {
      "--output_dir": "Export",
      "--layers": [
        "F.Cu",
        "In1.Cu",
        "In2.Cu",
        "In3.Cu",
        "In4.Cu",
        "In5.Cu",
        "In6.Cu",
        "In7.Cu",
        "In8.Cu",
        "B.Cu",
        "F.Paste",
        "B.Paste",
        "F.Silkscreen",
        "B.Silkscreen",
        "F.Mask",
        "B.Mask",
        "User.Drawings",
        "User.Comments",
        "Edge.Cuts",
        "F.Courtyard",
        "B.Courtyard",
        "F.Fab",
        "B.Fab"
      ],
      "kie_common_layers": [""],
      "--drawing-sheet": false,
      "--define-var": false,
      "--mirror": false,
      "--theme": "User",
      "--negative": false,
      "--black-and-white": false,
      "--sketch-pads-on-fab-layers": false,
      "--hide-DNP-footprints-on-fab-layers": false,
      "--sketch-DNP-footprints-on-fab-layers": false,
      "--crossout-DNP-footprints-on-fab-layers": false,
      "--page-size-mode": 0,
      "--fit-page-to-board": false,
      "--exclude-drawing-sheet": false,
      "--drill-shape-opt": 2,
      "--mode-single": false,
      "--mode-separate": false,
      "--mode-multipage": false,
      "--plot-invisible-text": false
    },
    "ddd": {
      "STEP": {
        "--output_dir": "Export",
        "--define-var": false,
        "--force": true,
        "--no-unspecified": false,
        "--no-dnp": false,
        "--grid-origin": false,
        "--drill-origin": false,
        "--subst-models": true,
        "--board-only": false,
        "--cut-vias-in-body": true,
        "--no-board-body": false,
        "--no-components": false,
        "--component-filter": false,
        "--include-tracks": true,
        "--include-pads": true,
        "--include-zones": true,
        "--include-inner-copper": false,
        "--include-silkscreen": true,
        "--include-soldermask": true,
        "--fuse-shapes": false,
        "--fill-all-vias": false,
        "--min-distance": false,
        "--net-filter": false,
        "--no-optimize-step": false,
        "--user-origin": false
      },
      "VRML": {
        "--output_dir": "Export",
        "--define-var": false,
        "--force": true,
        "--no-unspecified": false,
        "--no-dnp": false,
        "--user-origin": false,
        "--units": "mm",
        "--models-dir": false,
        "--models-relative": false
      }
    },
    "pcb_render": {
      "--output_dir": "Export",
      "--define-var": false,
      "PCB-Perspective": {
        "kie_type": "png",
        "kie_name_stub": "PCB-Perspective",
        "--width": 3200,
        "--height": 1800,
        "--side": "top",
        "--background": "default",
        "--quality": "basic",
        "--preset": false,
        "--floor": false,
        "--perspective": true,
        "--zoom ": 1,
        "--pan ": "0,0,0",
        "--pivot": "0,0,0",
        "--rotate": "-60,0,-45",
        "--light-top": false,
        "--light-bottom": false,
        "--light-side": false,
        "--light-camera": false,
        "--light-side-elevation": false
      },
      "PCB-Top": {
        "kie_type": "png",
        "kie_name_stub": "PCB-Top",
        "--width": 3200,
        "--height": 1800,
        "--side": "top",
        "--background": "default",
        "--quality": "basic",
        "--preset": false,
        "--floor": false,
        "--perspective": false,
        "--zoom ": 1,
        "--pan ": "0,0,0",
        "--pivot": "0,0,0",
        "--rotate": "0,0,-90",
        "--light-top": false,
        "--light-bottom": false,
        "--light-side": false,
        "--light-camera": false,
        "--light-side-elevation": false
      },
      "PCB-Bottom": {
        "kie_type": "png",
        "kie_name_stub": "PCB-Bottom",
        "--width": 3200,
        "--height": 1800,
        "--side": "bottom",
        "--background": "default",
        "--quality": "basic",
        "--preset": false,
        "--floor": false,
        "--perspective": false,
        "--zoom ": 1,
        "--pan ": "0,0,0",
        "--pivot": "0,0,0",
        "--rotate": "0,0,-90",
        "--light-top": false,
        "--light-bottom": false,
        "--light-side": false,
        "--light-camera": false,
        "--light-side-elevation": false
      },
      "PCB-Front": {
        "kie_type": "png",
        "kie_name_stub": "PCB-Front",
        "--width": 3200,
        "--height": 1800,
        "--side": "front",
        "--background": "default",
        "--quality": "basic",
        "--preset": false,
        "--floor": false,
        "--perspective": false,
        "--zoom ": 1,
        "--pan ": "0,0,0",
        "--pivot": "0,0,0",
        "--rotate": "0,0,-90",
        "--light-top": false,
        "--light-bottom": false,
        "--light-side": false,
        "--light-camera": false,
        "--light-side-elevation": false
      },
      "PCB-Back": {
        "kie_type": "png",
        "kie_name_stub": "PCB-Back",
        "--width": 3200,
        "--height": 1800,
        "--side": "back",
        "--background": "default",
        "--quality": "basic",
        "--preset": false,
        "--floor": false,
        "--perspective": false,
        "--zoom ": 1,
        "--pan ": "0,0,0",
        "--pivot": "0,0,0",
        "--rotate": "0,0,-90",
        "--light-top": false,
        "--light-bottom": false,
        "--light-side": false,
        "--light-camera": false,
        "--light-side-elevation": false
      },
      "PCB-Left": {
        "kie_type": "png",
        "kie_name_stub": "PCB-Left",
        "--width": 3200,
        "--height": 1800,
        "--side": "left",
        "--background": "default",
        "--quality": "basic",
        "--preset": false,
        "--floor": false,
        "--perspective": false,
        "--zoom ": 1,
        "--pan ": "0,0,0",
        "--pivot": "0,0,0",
        "--rotate": "0,0,-90",
        "--light-top": false,
        "--light-bottom": false,
        "--light-side": false,
        "--light-camera": false,
        "--light-side-elevation": false
      },
      "PCB-Right": {
        "kie_type": "png",
        "kie_name_stub": "PCB-Right",
        "--width": 3200,
        "--height": 1800,
        "--side": "right",
        "--background": "default",
        "--quality": "basic",
        "--preset": false,
        "--floor": false,
        "--perspective": false,
        "--zoom ": 1,
        "--pan ": "0,0,0",
        "--pivot": "0,0,0",
        "--rotate": "0,0,-90",
        "--light-top": false,
        "--light-bottom": false,
        "--light-side": false,
        "--light-camera": false,
        "--light-side-elevation": false
      }
    },
    "source_zip": {
      "--output_dir": "Output",
      "kie_ext_list": [
         ".kicad_pcb",
         ".kicad_sch",
         ".kicad_pro",
         ".kicad_prl",
         ".net"
      ]
    }
  }
}
'''

#=============================================================================================#

class Logger:
  def __init__ (self, log_file):
    self.terminal = sys.stdout
    self.buffer = []
    # self.log = open (log_file, "w", encoding = "utf-8")
    self.ansi_escape = re.compile (r'\x1B[@-_][0-?]*[ -/]*[@-~]')  # Regex to remove ANSI codes

    # Write timestamp at the beginning of the log session
    self.buffer.append (self._get_timestamp_header() + "\n")
  
  def _get_timestamp_header (self):
    now = datetime.now()
    offset = now.astimezone().strftime ('%z')  # e.g., +0530
    formatted_offset = offset [:3] + ':' + offset [3:]  # +05:30
    timestamp = now.strftime ('%I:%M:%S %p %d-%m-%Y, %A')
    return f"{formatted_offset} {timestamp}"

  def write (self, message):
    self.terminal.write (message)
    self.buffer.append (self.ansi_escape.sub ('', message))    # Strip them for the file

  def flush (self):
    # For compatibility with Python's standard streams
    self.terminal.flush()
    # self.log.flush()
  
  def save_to_file (self, filepath):
    self.buffer.append ("\n")

    # Get the version information from KiCad CLI and save that to the log file
    try:
      global current_config, default_config
      kicad_cli_path = f'{current_config.get ("kicad_cli_path", lambda: default_config ["kicad_cli_path"])}'
      # print (f"Running command: {color.blue (f'& \"{kicad_cli_path}\" version --format about')}")
      version_info = subprocess.run ([kicad_cli_path, "version", "--format", "about"], check = True, capture_output = True, text = True)
      self.buffer.append(f"\n{version_info.stdout.strip()}\n")
    except Exception as e:
      self.buffer.append (f"Error fetching KiCad CLI version: {e}\n")

    with open (filepath, 'w', encoding = 'utf-8') as f:
      f.writelines (self.buffer)
      print()
      print (f"KiExport [INFO]: Log saved to '{color.blue (filepath)}'")

#=============================================================================================#

class Colorize:
  def __init__(self, text):
    self.text = text
    self.ansi_code = '\033[0m'  # Default to reset

  def _color_text (self):
    return f"{self.ansi_code}{self.text}\033[0m"

  def red (self):
    self.ansi_code = '\033[31m'
    return self._color_text()

  def green (self):
    self.ansi_code = '\033[32m'
    return self._color_text()

  def yellow (self):
    self.ansi_code = '\033[33m'
    return self._color_text()

  def blue (self):
    self.ansi_code = '\033[34m'
    return self._color_text()

  def magenta (self):
    self.ansi_code = '\033[35m'
    return self._color_text()

  def cyan (self):
    self.ansi_code = '\033[36m'
    return self._color_text()

# Define ANSI escape codes for colors
COLORS = {
  'red': '\033[91m',
  'green': '\033[92m',
  'yellow': '\033[93m',
  'blue': '\033[94m',
  'magenta': '\033[95m',
  'cyan': '\033[96m',
  'reset': '\033[0m'
}
class _color:
  def __call__ (self, text, color):
    return f"{COLORS[color]}{text}{COLORS['reset']}"
  
  def __getattr__ (self, color):
    if color in COLORS:
      return lambda text: self(text, color)
    raise AttributeError (f"Color '{color}' is not supported.")

# Create an instance of the Colorize class
color = _color()

#=============================================================================================#

class LazyDict (dict):
  """
  A class to overload the get() function so that the value look up is only done when needed.

  Args:
      dict (dic): The dictionary to be converted.
  """
  def get (self, key, fallback = None):
    if key in self:
      return super().get (key)
    return fallback() if callable (fallback) else fallback

#=============================================================================================#

def to_lazy_dict (d):
  """
  Recursively convert all dicts to LazyDict.

  Args:
      d (dic): The normal dictionary to be converted.

  Returns:
      LazyDict : A dictionary of type LazyDict.
  """
  if isinstance (d, dict):
    return LazyDict ({k: to_lazy_dict (v) for k, v in d.items()})
  return d

#=============================================================================================#

def generateBomHtml (output_dir = None, pcb_filename = None, extra_args = None):
  """
  Runs the KiCad iBOM Python script on a specified PCB file.

  Args:
    pcb_filename (str): Path to the KiCad PCB file (.kicad_pcb).
    output_dir (str): Directory to save the output files. Defaults to the PCB file's directory.
    extra_args (list): Additional command-line arguments for customization (optional).

  Returns:
    str: Path to the generated iBOM HTML file.
  """

  # Read the paths.
  kicad_python_path = f'{current_config.get ("kicad_python_path", lambda: default_config ["kicad_python_path"])}'
  ibom_path = f'{current_config.get ("ibom_path", lambda: default_config ["ibom_path"])}'

  ibom_path = os.path.expandvars (ibom_path)

  # Check if the KiCad Python path exists.
  if not os.path.isfile (kicad_python_path):
    print (color.red (f"generateBomHtml [ERROR]: The KiCad Python path '{kicad_python_path}' does not exist. This command will be skipped."))
    command_exec_status ["bom_html"] = False
    return

  # Check if the iBOM script path exists.
  ibom_path = ibom_path.strip()
  ibom_path = os.path.normpath (ibom_path)
  if not os.path.isfile (ibom_path):
    print (color.red (f"generateBomHtml [ERROR]: The iBOM path '{ibom_path}' does not exist. This command will be skipped."))
    # st = os.stat (ibom_path)
    # print(f"Permissions: {oct(st.st_mode)}")
    command_exec_status ["bom_html"] = False
    return
  
  # Construct the iBOM command.
  ibom_export_command = [f'"{kicad_python_path}"', f'"{ibom_path}"']

  #---------------------------------------------------------------------------------------------#
  
  # Ensure PCB file exists.
  if not os.path.isfile (pcb_filename):
    print (color.red (f"generateBomHtml [ERROR]: The PCB file '{pcb_filename}' does not exist. The command will be skipped."))
    command_exec_status ["bom_html"] = False
    return
  
  #---------------------------------------------------------------------------------------------#

  file_name = extract_pcb_file_name (pcb_filename)
  file_name = file_name.replace (" ", "-") # If there are whitespace characters in the project name, replace them with a hyphen

  project_name = extract_project_name (file_name)
  info = extract_info_from_pcb (pcb_filename)
  
  print (f"generateBomHtml [INFO]: Project name is '{color.magenta (project_name)}' and revision is {color.magenta ('R')}{color.magenta (info ['rev'])}.")
  # ibom_filename = f"{project_name}-R{info ['rev']}-HTML-BoM-{filename_date}.html"

  #---------------------------------------------------------------------------------------------#

  file_path = os.path.abspath (pcb_filename) # Get the absolute path of the file.

  # Get the directory path of the file and save it as the project directory.
  # All other export directories will be relative to the project directory.
  project_dir = os.path.dirname (file_path)
  
  # Read the output directory name from the config file.
  od_from_config = project_dir + "/" + current_config.get ("data", {}).get ("bom", {}).get ("HTML").get ("--output_dir", lambda: default_config ["data"]["bom"]["HTML"]["--output_dir"])
  od_from_cli = output_dir  # The directory specified by the command line argument

  # Get the final directory path
  final_directory, filename_date = create_final_directory (od_from_config, od_from_cli, "BoM", info ["rev"], "generateBomHtml")

  #---------------------------------------------------------------------------------------------#

  # Form the initial part of the command
  
  full_command = []
  full_command.extend (ibom_export_command) # Add the base command
  
  seq_number = 1
  not_completed = True
  
  # Create the output file name.
  while not_completed:
    file_name = f"{final_directory}/{project_name}-R{info ['rev']}-BoM-HTML-{filename_date}-{seq_number}.html" # Extension needed here

    if os.path.exists (file_name):
      seq_number += 1
      not_completed = True
    else:
      full_command.append ("--dest-dir")
      full_command.append (f'"{final_directory}"') # Add the output file name with double quotes around it
      full_command.append ("--name-format")
      file_name = f"{project_name}-R{info ['rev']}-BoM-HTML-{filename_date}-{seq_number}" # No extension needed
      full_command.append (f'"{file_name}"')
      break
  
  #---------------------------------------------------------------------------------------------#

  # Get the argument list from the config file.
  arg_list = current_config.get ("data", {}).get ("bom", {}).get ("HTML")

  # Add the remaining arguments.
  # Check if the argument list is not an empty dictionary.
  if arg_list:
    for key, value in arg_list.items():
      if key.startswith ("--"): # Only fetch the arguments that start with "--"
        if key == "--output_dir": # Skip the --output_dir argument, sice we already added it
          continue

        elif key == "--name-format": # Skip the --name-format argument
          continue

        elif key == "--extra-data-file": # Check for the extra data file key
          full_command.append (key)
          full_command.append (f'"{file_path}"') # Add the PCB source file. This helps to populate the custom BoM fields.
        
        else:
          # Check if the value is empty
          if value == "": # Skip if the value is empty
            continue
          else:
            # Check if the vlaue is a JSON boolean
            if isinstance (value, bool):
              if value == True: # If the value is true, then append the key as an argument
                full_command.append (key)
            else:
              # Check if the value is a string and not a numeral
              if isinstance (value, str) and not value.isdigit():
                  full_command.append (key)
                  full_command.append (f'"{value}"') # Add as a double-quoted string
              elif isinstance (value, (int, float)):
                  full_command.append (key)
                  full_command.append (str (value))  # Append the numeric value as string

  # Finally add the input file
  full_command.append (f'"{pcb_filename}"')
  print ("generateBomHtml [INFO]: Running command: ", color.blue (' '.join (full_command)))

  #---------------------------------------------------------------------------------------------#

  # Run the iBOM script with error handling
  try:
    full_command = ' '.join (full_command) # Convert the list to a string
    subprocess.run (full_command, check = True)
    print (color.green (f"generateBomHtml [INFO]: Interactive HTML BoM generated successfully."))
    print()
    command_exec_status ["bom_html"] = True

  except subprocess.CalledProcessError as e:
    print (color.red (f"generateBomHtml [ERROR]: Error during HTML BoM generation: {e}"))
    print (color.red (f"generateBomHtml [INFO]: Make sure the 'Interactive HTML BoM' application is installed and available on the PATH."))
    print()
    command_exec_status ["bom_html"] = False

  except Exception as e:
    print (color.red (f" generateBomHtml [ERROR]: An unexpected error occurred: {e}"))
    print()
    command_exec_status ["bom_html"] = False

#=============================================================================================#

def merge_pdfs (folder_path, output_file, pdf_files = None):
  """
  Merges PDF files in a folder and creates a TOC based on file names.

  Args:
    folder_path (str): Path to the folder containing PDF files.
    output_file (str): Name of the output PDF file.
  """
  try:
    # List all PDF files in the specified folder.
    if pdf_files == None:
      pdf_files = [f for f in os.listdir (folder_path) if f.endswith ('.pdf')]
    
    if not pdf_files:
      print (f"merge_pdfs() [WARNING]: No PDF files found in the specified folder.")
      return

    # pdf_files.sort()  # Optional: sort the files alphabetically
    doc = pymupdf.open()  # Create a new PDF document
    toc = []  # List to hold the Table of Contents entries

    # Add each PDF to the document and create TOC entries
    for pdf in pdf_files:
      pdf_path = os.path.join (folder_path, pdf)
      try:
        with pymupdf.open (pdf_path) as pdf_doc:
          start_page = doc.page_count  # Get the starting page number
          doc.insert_pdf (pdf_doc)  # Merge the PDF
          toc.append ((1, pdf [:-4], start_page + 1))  # Add TOC entry

      except Exception as e:
        print (color.red (f"merge_pdfs() [ERROR]: Error processing file {pdf}: {e}"))

    # Set the Table of Contents
    doc.set_toc (toc)

    # Save the merged document
    output_path = os.path.join (folder_path, output_file)
    doc.save (output_path)
    doc.close()

    # Delete original PDF files.
    for pdf in pdf_files:
      os.remove (os.path.join (folder_path, pdf))

    print (f"Merged PDF created: {output_path}")
    print ("Original PDF files have been deleted.")

  except PermissionError:
    print (color.red ("merge_pdfs() [ERROR]: Unable to access the specified folder or files."))
  except Exception as e:
    print (color.red (f"merge_pdfs() [ERROR]: {e}"))

#=============================================================================================#

def generateGerbers (output_dir, pcb_filename, to_overwrite = True):
  # Generate the drill files first if specified
  kie_include_drill = current_config.get ("data", {}).get ("gerbers", {}).get ("kie_include_drill", lambda: default_config ["data"]["gerbers"]["kie_include_drill"])

  # Get the KiCad CLI path.
  kicad_cli_path = f'{current_config.get ("kicad_cli_path", lambda: default_config ["kicad_cli_path"])}'

  # Check if the value is boolean and then true or false
  if isinstance (kie_include_drill, bool):
    kie_include_drill = str (kie_include_drill).lower()

  if kie_include_drill == "true":
    kie_include_drill = True
  elif kie_include_drill == "false":
    kie_include_drill = False

  if kie_include_drill == True:
    generateDrills (output_dir, pcb_filename)
  
  #---------------------------------------------------------------------------------------------#
  
  # Common base command
  gerber_export_command = [f'"{kicad_cli_path}"', "pcb", "export", "gerbers"]

  # Check if the pcb file exists
  if not check_file_exists (pcb_filename):
    print (color.red (f"generateGerbers [ERROR]: '{pcb_filename}' does not exist."))
    command_exec_status ["gerbers"] = False
    return

  #---------------------------------------------------------------------------------------------#
  
  file_name = extract_pcb_file_name (pcb_filename)
  file_name = file_name.replace (" ", "-") # If there are whitespace characters in the project name, replace them with a hyphen
  
  project_name = extract_project_name (file_name)
  info = extract_info_from_pcb (pcb_filename)
  print (f"generateGerbers [INFO]: Project name is '{color.magenta (project_name)}' and revision is {color.magenta ('R')}{color.magenta (info ['rev'])}.")
  
  #---------------------------------------------------------------------------------------------#
  
  file_path = os.path.abspath (pcb_filename) # Get the absolute path of the file.

  # Get the directory path of the file and save it as the project directory.
  # All other export directories will be relative to the project directory.
  project_dir = os.path.dirname (file_path)
  
  # Read the output directory name from the config file.
  od_from_config = project_dir + "/" + current_config.get ("data", {}).get ("gerbers", {}).get ("--output_dir", lambda: default_config ["data"]["gerbers"]["--output_dir"])
  od_from_cli = output_dir  # The output directory specified by the command line argument

  # Get the final directory path.
  final_directory, filename_date = create_final_directory (od_from_config, od_from_cli, "Gerber", info ["rev"], "generateGerbers")

  #---------------------------------------------------------------------------------------------#
  
  # Get the argument list from the config file.
  arg_list = current_config.get ("data", {}).get ("gerbers", {})

  seq_number = 1
  not_completed = True
  full_command = []
  full_command.extend (gerber_export_command) # Add the base command
  full_command.append ("--output")
  full_command.append (f'"{final_directory}"')
  
  # Add the remaining arguments.
  # Check if the argument list is not an empty dictionary.
  if arg_list:
    for key, value in arg_list.items():
      if key.startswith ("--"): # Only fetch the arguments that start with "--"
        if key == "--output_dir": # Skip the --output_dir argument, sice we already added it
          continue
        elif key == "--layers":
          full_command.append (key)
          layers_csv = ",".join (value) # Convert the list to a comma-separated string
          full_command.append (f'"{layers_csv}"')
        else:
          # Check if the value is empty
          if value == "": # Skip if the value is empty
            continue
          else:
            # Check if the vlaue is a JSON boolean
            if isinstance (value, bool):
              if value == True: # If the value is true, then append the key as an argument
                full_command.append (key)
            else:
              # Check if the value is a string and not a numeral
              if isinstance (value, str) and not value.isdigit():
                  full_command.append (key)
                  full_command.append (f'"{value}"') # Add as a double-quoted string
              elif isinstance (value, (int, float)):
                  full_command.append (key)
                  full_command.append (str (value))  # Append the numeric value as string
  
  # Finally add the input file
  full_command.append (f'"{pcb_filename}"')
  print ("generateGerbers [INFO]: Running command: ", color.blue (' '.join (full_command)))
  
  #---------------------------------------------------------------------------------------------#
  
  # Delete the existing files in the output directory
  delete_files (final_directory, include_extensions = [".gbr", ".gbrjob"])
  
  #---------------------------------------------------------------------------------------------#
  
  # Run the command
  try:
    full_command = ' '.join (full_command) # Convert the list to a string
    subprocess.run (full_command, check = True)
    print (color.green ("generateGerbers [OK]: Gerber files exported successfully."))
    command_exec_status ["gerbers"] = True
  
  except subprocess.CalledProcessError as e:
    print (color.red (f"generateGerbers [ERROR]: Error occurred: {e}"))
    command_exec_status ["gerbers"] = False
    return
  
  #---------------------------------------------------------------------------------------------#
  
  # Rename the files by adding Revision after the project name.
  rename_files (final_directory, project_name, info ['rev'], [".gbr", ".gbrjob"])
  
  #---------------------------------------------------------------------------------------------#
  
  #---------------------------------------------------------------------------------------------#
  
  # Due to above rename of project files, MUST also rename the project files within the gbrjob file.
  # The gbrjob file is actually a json format file so easily done with json load and save
  gbrjob_file_path=os.path.join(final_directory,project_name+"-R"+info ['rev']+"-job.gbrjob")

  # Function to modify the Path value
  def modify_path(original_path,prefix,revision):
    base_name = original_path [len (prefix):]  # Remove the prefix part
    return f"{prefix}-R{revision}{base_name}"

  # Load the gbrjob (JSON) file
  with open(gbrjob_file_path, "r", encoding="utf-8") as f:
    data = json.load(f)

  # Modify all Path values inside FilesAttributes
  if "FilesAttributes" in data:
    for file_attr in data["FilesAttributes"]:
      if "Path" in file_attr:
        file_attr["Path"] = modify_path(file_attr["Path"],project_name,info ['rev'])

    # Save the modified JSON back to file
  with open(gbrjob_file_path, "w", encoding="utf-8") as f:
    json.dump(data, f, indent=2, ensure_ascii=False)

  #---------------------------------------------------------------------------------------------#
  
  seq_number = 1
  not_completed = True

  #files_to_include = [".gbr", ".gbrjob"]
  # Do not include the gbrjob file in the output zip for mfc
  files_to_include = [".gbr"]

  if kie_include_drill:
    files_to_include.extend ([".drl", ".ps", ".pdf"])
  
  # Sequentially name and create the zip files.
  while not_completed:
    zip_file_name = f"{project_name}-R{info ['rev']}-Gerber-{filename_date}-{seq_number}.zip"

    if os.path.exists (f"{final_directory}/{zip_file_name}"):
      seq_number += 1
    else:
      zip_cnt = zip_all_files_ext (final_directory, zip_file_name, files_to_include)
      print (f"generateGerbers [OK]: ZIP file '{color.magenta (zip_file_name)}' created successfully with {zip_cnt} files.")
      print()
      not_completed = False

#=============================================================================================#

def generateDrills (output_dir, pcb_filename):
  # Get the KiCad CLI path.
  kicad_cli_path = f'{current_config.get ("kicad_cli_path", lambda: default_config ["kicad_cli_path"])}'

  # Common base command
  drill_export_command = [f'"{kicad_cli_path}"', "pcb", "export", "drill"]

  # Check if the pcb file exists
  if not check_file_exists (pcb_filename):
    print (color.red (f"generateDrills [ERROR]: '{pcb_filename}' does not exist."))
    command_exec_status ["drills"] = False
    return

  #-------------------------------------------------------------------------------------------#
  
  file_name = extract_pcb_file_name (pcb_filename)
  file_name = file_name.replace (" ", "-") # If there are whitespace characters in the project name, replace them with a hyphen
  
  project_name = extract_project_name (file_name)
  info = extract_info_from_pcb (pcb_filename)
  print (f"generateDrills [INFO]: Project name is '{color.magenta (project_name)}' and revision is {color.magenta ('R')}{color.magenta (info ['rev'])}.")
  
  #-------------------------------------------------------------------------------------------#

  file_path = os.path.abspath (pcb_filename) # Get the absolute path of the file.

  # Get the directory path of the file and save it as the project directory.
  # All other export directories will be relative to the project directory.
  project_dir = os.path.dirname (file_path)
  
  # Read the target directory name from the config file
  od_from_config = project_dir + "/" + current_config.get ("data", {}).get ("drills", {}).get ("--output_dir", lambda: default_config ["data"]["drills"]["--output_dir"])
  od_from_cli = output_dir  # The directory specified by the command line argument

  # Get the final directory path
  final_directory, filename_date = create_final_directory (od_from_config, od_from_cli, "Gerber", info ["rev"], "generateDrills")

  # Check if the final directory ends with a slash, and add one if not.
  if final_directory [-1] != "/":
    final_directory = f"{final_directory}/"
    
  #-------------------------------------------------------------------------------------------#
  
  # Get the argument list from the config file.
  arg_list = current_config.get ("data", {}).get ("drills", {})

  seq_number = 1
  not_completed = True
  full_command = []
  full_command.extend (drill_export_command) # Add the base command
  full_command.append ("--output")
  full_command.append (f'"{final_directory}"')
  
  # Add the remaining arguments.
  # Check if the argument list is not an empty dictionary.
  if arg_list:
    for key, value in arg_list.items():
      if key.startswith ("--"): # Only fetch the arguments that start with "--"
        if key == "--output_dir": # Skip the --output_dir argument, sice we already added it
          continue
        else:
          # Check if the value is empty
          if value == "": # Skip if the value is empty
            continue
          else:
            # Check if the vlaue is a JSON boolean
            if isinstance (value, bool):
              if value == True: # If the value is true, then append the key as an argument
                full_command.append (key)
            else:
              # Check if the value is a string and not a numeral
              if isinstance (value, str) and not value.isdigit():
                  full_command.append (key)
                  full_command.append (f'"{value}"') # Add as a double-quoted string
              elif isinstance (value, (int, float)):
                  full_command.append (key)
                  full_command.append (str (value))  # Append the numeric value as string
  
  # Finally add the input file
  full_command.append (f'"{pcb_filename}"')
  print ("generateDrills [INFO]: Running command: ", color.blue (' '.join (full_command)))
  
  #-------------------------------------------------------------------------------------------#

  # Delete the existing files in the output directory
  delete_files (final_directory, include_extensions = [".drl", ".ps", ".pdf"])

  #-------------------------------------------------------------------------------------------#
  
  # Run the command
  try:
    full_command = ' '.join (full_command) # Convert the list to a string
    subprocess.run (full_command, check = True)
    print (color.green ("generateDrills [OK]: Drill files exported successfully."))
    print()
    command_exec_status ["drills"] = True
  
  except subprocess.CalledProcessError as e:
    print (color.red (f"generateDrills [ERROR]: Error occurred: {e}"))
    print()
    command_exec_status ["drills"] = False
    return
  
  #-------------------------------------------------------------------------------------------#

  # Rename the files by adding Revision after the project name.
  rename_files (final_directory, project_name, info ['rev'], [".drl", ".ps", ".pdf"])

#=============================================================================================#

def generatePositions (output_dir, pcb_filename, to_overwrite = True):
  global current_config  # Access the global config
  global default_config  # Access the global config
  
  # Get the KiCad CLI path.
  kicad_cli_path = f'{current_config.get ("kicad_cli_path", lambda: default_config ["kicad_cli_path"])}'

  # Common base command
  position_export_command = [f'"{kicad_cli_path}"', "pcb", "export", "pos"]

  # Check if the input file exists
  if not check_file_exists (pcb_filename):
    print (color.red (f"generatePositions [ERROR]: '{pcb_filename}' does not exist."))
    command_exec_status ["positions"] = False
    return

  #---------------------------------------------------------------------------------------------#
  
  file_name = extract_pcb_file_name (pcb_filename)
  file_name = file_name.replace (" ", "-") # If there are whitespace characters in the project name, replace them with a hyphen
  
  project_name = extract_project_name (file_name)
  info = extract_info_from_pcb (pcb_filename)
  print (f"generatePositions [INFO]: Project name is '{color.magenta (project_name)}' and revision is {color.magenta ('R')}{color.magenta (info ['rev'])}.")
  
  #---------------------------------------------------------------------------------------------#

  file_path = os.path.abspath (pcb_filename) # Get the absolute path of the file.

  # Get the directory path of the file and save it as the project directory.
  # All other export directories will be relative to the project directory.
  project_dir = os.path.dirname (file_path)
  
  # Read the output directory name from the config file.
  od_from_config = project_dir + "/" + current_config.get ("data", {}).get ("positions", {}).get ("--output_dir", lambda: default_config ["data"]["positions"]["--output_dir"])
  od_from_cli = output_dir  # The directory specified by the command line argument

  # Get the final directory path
  final_directory, filename_date = create_final_directory (od_from_config, od_from_cli, "Position", info ["rev"], "generatePositions")
  
  #---------------------------------------------------------------------------------------------#
  
  pos_front_filename = f"{final_directory}/{project_name}-Pos-Front.csv"
  pos_back_filename = f"{final_directory}/{project_name}-Pos-Back.csv"
  pos_all_filename = f"{final_directory}/{project_name}-Pos-All.csv"

  # Create a list of filenames for front, back, and both.
  pos_filenames = [pos_front_filename, pos_back_filename, pos_all_filename]

  # Create a list of three command sets for front, back, and both.
  full_command_list = []
  for filename in pos_filenames:
      full_command = position_export_command.copy()  # Copy the base command
      full_command.append ("--output")
      full_command.append (f'"{filename}"')
      full_command_list.append (full_command)
  
  # Get the argument list from the config file.
  arg_list = current_config.get ("data", {}).get ("positions", {})
  sides = arg_list.get ("--side", None) # Get the sides from the config file as a string

  # Check if the sides are valid and apply the default value if not
  if sides == None or sides == "":
    print (color.yellow (f"generatePositions [INFO]: No sides specified. Using both sides."))
    sides = "both"
  
  # Add the remaining arguments.
  # Check if the argument list is not an empty dictionary.
  if arg_list:
    for i, command_set in enumerate (full_command_list):
      for key, value in arg_list.items():
        if key.startswith ("--"): # Only fetch the arguments that start with "--"
          if key == "--output_dir": # Skip the --output_dir argument, sice we already added it
            continue
          elif key == "--side":
            if sides.__contains__ ("front") and i == 0:
              command_set.append (key)
              command_set.append ("front")
            elif sides.__contains__ ("back") and i == 1:
              command_set.append (key)
              command_set.append ("back")
            elif sides.__contains__ ("both") and i == 2:
              command_set.append (key)
              command_set.append ("both")
          else:
            # Check if the value is empty
            if value == "": # Skip if the value is empty
              continue
            else:
              # Check if the vlaue is a JSON boolean
              if isinstance (value, bool):
                if value == True: # If the value is true, then append the key as an argument
                  command_set.append (key)
              else:
                # Check if the value is a string and not a numeral
                if isinstance (value, str) and not value.isdigit():
                    command_set.append (key)
                    command_set.append (f'"{value}"') # Add as a double-quoted string
                elif isinstance (value, (int, float)):
                    command_set.append (key)
                    command_set.append (str (value))  # Append the numeric value as string

  # Finally append the filename to the commands
  for command_set in full_command_list:
    # board_file_path = os.path.abspath (pcb_filename)
    command_set.append (f'"{pcb_filename}"')
  
  #---------------------------------------------------------------------------------------------#
  
  # Delete all non-zip files
  delete_non_zip_files (final_directory)
  
  #---------------------------------------------------------------------------------------------#
  
  # Run the commands
  for i, full_command in enumerate (full_command_list):
    if (sides.__contains__ ("front") and i == 0) or (sides.__contains__ ("back") and i == 1) or (sides.__contains__ ("both") and i == 2):
      try:
        command_string = ' '.join (full_command)  # Convert the list to a string
        print (f"generatePositions [INFO]: Running command: {color.blue (command_string)}")
        subprocess.run (command_string, check = True)
      except subprocess.CalledProcessError as e:
        print (color.red (f"generatePositions [ERROR]: Error occurred while generating the files."))
        command_exec_status ["positions"] = False
        return

  print (color.green ("generatePositions [OK]: Position files exported successfully."))
  command_exec_status ["positions"] = True
  
  #---------------------------------------------------------------------------------------------#
  
  # Rename the files by adding Revision after the project name.
  rename_files (final_directory, project_name, info ['rev'], [".csv"])
  
  #---------------------------------------------------------------------------------------------#
  
  seq_number = 1
  not_completed = True

  files_to_include = [".csv"]
  
  # Sequentially name and create the zip files.
  while not_completed:
    zip_file_name = f"{project_name}-R{info ['rev']}-Position-Files-{filename_date}-{seq_number}.zip"

    if os.path.exists (f"{final_directory}/{zip_file_name}"):
      seq_number += 1
    else:
      zip_cnt = zip_all_files_ext (final_directory, zip_file_name, files_to_include)
      print (f"generatePositions [OK]: ZIP file '{color.magenta (zip_file_name)}' created successfully with {zip_cnt} files.")
      print()
      not_completed = False

#=============================================================================================#

def generatePcbPdf (output_dir, pcb_filename, to_overwrite = True):
  # Get the KiCad CLI path.
  kicad_cli_path = f'{current_config.get ("kicad_cli_path", lambda: default_config ["kicad_cli_path"])}'

  # Common base command
  pcb_pdf_export_command = [f'"{kicad_cli_path}"', "pcb", "export", "pdf"]

  # Check if the pcb file exists
  if not check_file_exists (pcb_filename):
    print (color.red (f"generatePcbPdf [ERROR]: '{pcb_filename}' does not exist."))
    command_exec_status ["pcb_pdf"] = False
    return

  #---------------------------------------------------------------------------------------------#

  file_name = extract_pcb_file_name (pcb_filename)
  file_name = file_name.replace (" ", "-") # If there are whitespace characters in the project name, replace them with a hyphen
  
  project_name = extract_project_name (file_name)
  info = extract_info_from_pcb (pcb_filename)
  print (f"generatePcbPdf [INFO]: Project name is '{color.magenta (project_name)}' and revision is {color.magenta ('R')}{color.magenta (info ['rev'])}.")
  
  #---------------------------------------------------------------------------------------------#

  file_path = os.path.abspath (pcb_filename) # Get the absolute path of the file.

  # Get the directory path of the file and save it as the project directory.
  # All other export directories will be relative to the project directory.
  project_dir = os.path.dirname (file_path)
  
  # Read the output directory name from the config file.
  od_from_config = project_dir + "/" + current_config.get ("data", {}).get ("pcb_pdf", {}).get ("--output_dir", lambda: default_config ["data"]["pcb_pdf"]["--output_dir"])
  od_from_cli = output_dir  # The output directory specified by the command line argument

  # Get the final directory path
  final_directory, filename_date = create_final_directory (od_from_config, od_from_cli, "PCB", info ["rev"], "generatePcbPdf")

  #---------------------------------------------------------------------------------------------#
  
  # Delete the existing files in the output directory
  delete_files (final_directory, include_extensions = [".pdf", ".ps"])

  #---------------------------------------------------------------------------------------------#
  
  # Get the argument list from the config file.
  arg_list = current_config.get ("data", {}).get ("pcb_pdf", {})

  # Check the number of technical layers to export. This is not the number of copper layers.
  layer_count = len (arg_list.get ("--layers", []))

  if layer_count <= 0:
    print (color.red (f"generatePcbPdf [ERROR]: No layers specified for export."))
    command_exec_status ["pcb_pdf"] = False
    return

  # Get the number of common layers to include in each of the PDF.
  # common_layer_count = len (arg_list.get ("kie_common_layers", []))

  seq_number = 1
  not_completed = True
  base_command = []
  base_command.extend (pcb_pdf_export_command) # Add the base command
  pdf_files = []      # List of generated PDF files in order of Layers specified by user      
  
  for i in range (layer_count):
    full_command = base_command [:]
    # Get the arguments.
    if arg_list: # Check if the argument list is not an empty dictionary.
      for key, value in arg_list.items():
        if key.startswith ("--"): # Only fetch the arguments that start with "--"
          if key == "--output_dir": # Skip the --output_dir argument, sice we already added it
            continue
          elif key == "--layers":
            layer_name = arg_list ["--layers"][i] # Get a layer name from the layer list
            layer_name = layer_name.replace (".", "_") # Replace dots with underscores
            layer_name = layer_name.replace (" ", "_") # Replace spaces with underscores

            full_command.append ("--output")
            pdf_file_name = f'"{final_directory}/{project_name}-R{info ["rev"]}-{layer_name}.pdf"' # This is the ouput file name, and not a directory name
            full_command.append (pdf_file_name)
            pdf_file_name = f"{project_name}-R{info ["rev"]}-{layer_name}.pdf" # This is the ouput file name, and not a directory name
            pdf_files.append(os.path.join(final_directory, pdf_file_name))

            layer_name = arg_list ["--layers"][i] # Get a layer name from the layer list
            layer_list = [f"{layer_name}"]  # Now create a list with the first item as the layer name
            common_layer_list = arg_list ["kie_common_layers"]  # Add the common layers
            layer_list.extend (common_layer_list) # Now combine the two lists
            layers_csv = ",".join (layer_list) # Convert the list to a comma-separated string
            full_command.append (key)
            full_command.append (f'"{layers_csv}"')
          else:
            # Check if the value is empty
            if value == "": # Skip if the value is empty
              continue
            else:
              # Check if the vlaue is a JSON boolean
              if isinstance (value, bool):
                # Special handline for --mirror flag for Fab layers generation
                if key == "--mirror":
                  if layer_name == "F.Fab" or layer_name == "F.Silkscreen" or layer_name == "F.Cu" or layer_name == "F.Mask" or layer_name == "F.Paste" or layer_name == "F.Courtyard": # Skip mirror options for all Front layers
                    continue
                  elif layer_name == "B.Fab" or layer_name == "B.Silkscreen" or layer_name == "B.Cu" or layer_name == "B.Mask" or layer_name == "B.Paste" or layer_name == "B.Courtyard": # Force mirror options for all Bottom layers
                    full_command.append (key)
                  elif value == True: # If the value is true, then append the key as an argument
                    full_command.append (key)
                elif value == True: # If the value is true, then append the key as an argument
                  full_command.append (key)
              else:
                # Check if the value is a string and not a numeral
                if isinstance (value, str) and not value.isdigit():
                  full_command.append (key)
                  full_command.append (f'"{value}"') # Add as a double-quoted string
                elif isinstance (value, (int, float)):
                  # Special handline for --drill-shape-opt value for silk/Fab layers generation
                  if key == "--drill-shape-opt":
                    if layer_name == "F.Silkscreen" or layer_name == "B.Silkscreen" or layer_name == "F.Fab" or layer_name == "B.Fab": # Change drill_shape_opt options for Fab and Silk layers
                      full_command.append (key)
                      full_command.append (str (0))  # Force drill marking as None
                    else:
                      full_command.append (key)
                      full_command.append (str (value))  # Append the numeric value as string
                  else:
                    full_command.append (key)
                    full_command.append (str (value))  # Append the numeric value as string

    full_command.append (f'"{pcb_filename}"')
    print ("generatePcbPdf [INFO]: Running command: ", color.blue (' '.join (full_command)))

      # Run the command
    try:
      full_command = ' '.join (full_command) # Convert the list to a string
      subprocess.run (full_command, check = True)
      # print (color.green ("generatePcbPdf [OK]: PCB PDF files exported successfully."))
      command_exec_status ["pcb_pdf"] = True
    
    except subprocess.CalledProcessError as e:
      print (color.red (f"generatePcbPdf [ERROR]: Error occurred: {e}"))
      command_exec_status ["pcb_pdf"] = True
      continue
  
  #---------------------------------------------------------------------------------------------#
  
  # # Generate a single file if specified
  # kie_single_file = current_config.get ("data", {}).get ("pcb_pdf", {}).get ("kie_single_file", lambda: default_config ["data"]["pcb_pdf"]["kie_single_file"])

  # # Check if the value is boolean and then true or false
  # if isinstance (kie_single_file, bool):
  #   kie_single_file = str (kie_single_file).lower()

  # if kie_single_file == "true":
  #   kie_single_file = True
  # elif kie_single_file == "false":
  #   kie_single_file = False

  # if kie_single_file == True:
  #   full_command.append (f'"{pcb_filename}"')
  
  #---------------------------------------------------------------------------------------------#

  # Merge all the PDFs into one file
  merged_pdf_filename = f"{project_name}-R{info ['rev']}-PCB-PDF-All-{filename_date}-{seq_number}.pdf"
  merge_pdfs (final_directory, merged_pdf_filename, pdf_files)

  #---------------------------------------------------------------------------------------------#

  print (color.green ("generatePcbPdf [OK]: PCB PDF files exported successfully."))

  #---------------------------------------------------------------------------------------------#

  # Generating a single pdf file no need for zip all (all single files have been deleted already as well)
  #seq_number = 1
  #not_completed = True

  #files_to_include = [".pdf"]
  
  # Sequentially name and create the zip files.
  #while not_completed:
  #  zip_file_name = f"{project_name}-R{info ['rev']}-PCB-PDF-{filename_date}-{seq_number}.zip"

  #  if os.path.exists (f"{final_directory}/{zip_file_name}"):
  #    seq_number += 1
  #  else:
  #    zip_cnt = zip_all_files_ext (final_directory, zip_file_name, files_to_include)
  #    print (f"generatePcbPdf [OK]: ZIP file '{color.magenta (zip_file_name)}' created successfully with {zip_cnt} files.")
  #    print()
  #    not_completed = False

#==============================================================================================#

def generatePcbFabPdf (output_dir, pcb_filename, to_overwrite = True):
  # Get the KiCad CLI path.
  kicad_cli_path = f'{current_config.get ("kicad_cli_path", lambda: default_config ["kicad_cli_path"])}'

  # Common base command
  pcb_pdf_export_command = [f'"{kicad_cli_path}"', "pcb", "export", "pdf"]

  # Check if the pcb file exists
  if not check_file_exists (pcb_filename):
    print (color.red (f"generatePcbFabPdf [ERROR]: '{pcb_filename}' does not exist."))
    command_exec_status ["pcb_fab_pdf"] = False
    return

  #---------------------------------------------------------------------------------------------#

  file_name = extract_pcb_file_name (pcb_filename)
  file_name = file_name.replace (" ", "-") # If there are whitespace characters in the project name, replace them with a hyphen
  
  project_name = extract_project_name (file_name)
  info = extract_info_from_pcb (pcb_filename)
  print (f"generatePcbFabPdf [INFO]: Project name is '{color.magenta (project_name)}' and revision is {color.magenta ('R')}{color.magenta (info ['rev'])}.")
  
  #---------------------------------------------------------------------------------------------#

  file_path = os.path.abspath (pcb_filename) # Get the absolute path of the file.

  # Get the directory path of the file and save it as the project directory.
  # All other export directories will be relative to the project directory.
  project_dir = os.path.dirname (file_path)
  
  # Read the output directory name from the config file.
  od_from_config = project_dir + "/" + current_config.get ("data", {}).get ("pcb_fab_pdf", {}).get ("--output_dir", lambda: default_config ["data"]["pcb_fab_pdf"]["--output_dir"])
  od_from_cli = output_dir  # The output directory specified by the command line argument

  # Get the final directory path
  final_directory, filename_date = create_final_directory (od_from_config, od_from_cli, "PCB-Fab", info ["rev"], "generatePcbFabPdf")

  #---------------------------------------------------------------------------------------------#
  
  # Delete the existing files in the output directory
  delete_files (final_directory, include_extensions = [".pdf", ".ps"])

  #---------------------------------------------------------------------------------------------#
  
  # Get the argument list from the config file.
  arg_list = current_config.get ("data", {}).get ("pcb_fab_pdf", {})

  # Check the number of technical layers to export. This is not the number of copper layers.
  layer_count = len (arg_list.get ("--layers", []))

  if layer_count <= 0:
    print (color.red (f"generatePcbFabPdf [ERROR]: No layers specified for export."))
    command_exec_status ["pcb_fab_pdf"] = False
    return

  # Get the number of common layers to include in each of the PDF.
  # common_layer_count = len (arg_list.get ("kie_common_layers", []))

  seq_number = 1
  not_completed = True
  base_command = []
  base_command.extend (pcb_pdf_export_command) # Add the base command
  pdf_files = []      # List of generated PDF files in order of Layers specified by user      
  
  for i in range (layer_count):
    full_command = base_command [:]
    # Get the arguments.
    if arg_list: # Check if the argument list is not an empty dictionary.
      for key, value in arg_list.items():
        if key.startswith ("--"): # Only fetch the arguments that start with "--"
          if key == "--output_dir": # Skip the --output_dir argument, since we already added it
            continue
          elif key == "--layers":
            layer_name = arg_list ["--layers"][i] # Get a layer name from the layer list
            layer_name = layer_name.replace (".", "_") # Replace dots with underscores
            layer_name = layer_name.replace (" ", "_") # Replace spaces with underscores

            full_command.append ("--output")
            pdf_file_name = f'"{final_directory}/{project_name}-R{info ["rev"]}-{layer_name}.pdf"' # This is the ouput file name, and not a directory name
            full_command.append (pdf_file_name)
            pdf_file_name = f"{project_name}-R{info ["rev"]}-{layer_name}.pdf" # This is the ouput file name, and not a directory name
            pdf_files.append(os.path.join(final_directory, pdf_file_name))

            layer_name = arg_list ["--layers"][i] # Get a layer name from the layer list
            layer_list = [f"{layer_name}"]  # Now create a list with the first item as the layer name
            common_layer_list = arg_list ["kie_common_layers"]  # Add the common layers
            layer_list.extend (common_layer_list) # Now combine the two lists
            layers_csv = ",".join (layer_list) # Convert the list to a comma-separated string
            full_command.append (key)
            full_command.append (f'"{layers_csv}"')
          else:
            # Check if the value is empty
            if value == "": # Skip if the value is empty
              continue
            else:
              # Check if the vlaue is a JSON boolean
              if isinstance (value, bool):
                # Special handline for --mirror flag for Fab layers generation
                if key == "--mirror":
                  if layer_name == "F.Fab": # Skip mirror options for F.Fab layer
                    continue
                  elif layer_name == "B.Fab": # Force mirror options for B.Fab layer
                    full_command.append (key)
                  elif value == True: # If the value is true, then append the key as an argument
                    full_command.append (key)
                elif value == True: # If the value is true, then append the key as an argument
                  full_command.append (key)
              else:
                # Check if the value is a string and not a numeral
                if isinstance (value, str) and not value.isdigit():
                  full_command.append (key)
                  full_command.append (f'"{value}"') # Add as a double-quoted string
                elif isinstance (value, (int, float)):
                  # Special handline for --drill-shape-opt value for silk/Fab layers generation
                  if key == "--drill-shape-opt":
                    if layer_name == "F.Silkscreen" or layer_name == "B.Silkscreen" or layer_name == "F.Fab" or layer_name == "B.Fab": # Skip mirror options for F.Fab layer
                      full_command.append (key)
                      full_command.append (str (0))  # Force drill marking as None
                    else:
                      full_command.append (key)
                      full_command.append (str (value))  # Append the numeric value as string
                  else:
                    full_command.append (key)
                    full_command.append (str (value))  # Append the numeric value as string

    full_command.append (f'"{pcb_filename}"')
    print ("generatePcbFabPdf [INFO]: Running command: ", color.blue (' '.join (full_command)))

      # Run the command
    try:
      full_command = ' '.join (full_command) # Convert the list to a string
      subprocess.run (full_command, check = True)
      # print (color.green ("generatePcbFabPdf [OK]: PCB Fab PDF files exported successfully."))
      command_exec_status ["pcb_fab_pdf"] = True
    
    except subprocess.CalledProcessError as e:
      print (color.red (f"generatePcbFabPdf [ERROR]: Error occurred: {e}"))
      command_exec_status ["pcb_fab_pdf"] = True
      continue
  
  #---------------------------------------------------------------------------------------------#
  
  # # Generate a single file if specified
  # kie_single_file = current_config.get ("data", {}).get ("pcb_pdf", {}).get ("kie_single_file", lambda: default_config ["data"]["pcb_pdf"]["kie_single_file"])

  # # Check if the value is boolean and then true or false
  # if isinstance (kie_single_file, bool):
  #   kie_single_file = str (kie_single_file).lower()

  # if kie_single_file == "true":
  #   kie_single_file = True
  # elif kie_single_file == "false":
  #   kie_single_file = False

  # if kie_single_file == True:
  #   full_command.append (f'"{pcb_filename}"')
  
  #---------------------------------------------------------------------------------------------#

  # Merge all the PDFs into one file
  merged_pdf_filename = f"{project_name}-R{info ['rev']}-PCB-PDF-FAB-{filename_date}-{seq_number}.pdf"
  merge_pdfs (final_directory, merged_pdf_filename, pdf_files)

  #---------------------------------------------------------------------------------------------#

  print (color.green ("generatePcbFabPdf [OK]: PCB Fab PDF files exported successfully."))

  #---------------------------------------------------------------------------------------------#

  # Generating a single pdf file no need for zip all
  #seq_number = 1
  #not_completed = True

  #files_to_include = [".pdf"]
  
  # Sequentially name and create the zip files.
  #while not_completed:
  #  zip_file_name = f"{project_name}-R{info ['rev']}-PCB-PDF-{filename_date}-{seq_number}.zip"

  #  if os.path.exists (f"{final_directory}/{zip_file_name}"):
  #    seq_number += 1
  #  else:
  #    zip_cnt = zip_all_files_ext (final_directory, zip_file_name, files_to_include)
  #    print (f"generatePcbFabPdf [OK]: ZIP file '{color.magenta (zip_file_name)}' created successfully with {zip_cnt} files.")
  #    print()
  #    not_completed = False

#==============================================================================================#

def generatePcbRenders (output_dir, pcb_filename, preset = None, to_overwrite = True):
  global current_config  # Access the global config
  global default_config  # Access the global config
  
  # Get the KiCad CLI path.
  kicad_cli_path = f'{current_config.get ("kicad_cli_path", lambda: default_config ["kicad_cli_path"])}'

  # Common base command.
  pcb_render_export_command = [f'"{kicad_cli_path}"', "pcb", "render"]

  # Check if the input file exists.
  if not check_file_exists (pcb_filename):
    print (color.red (f"generatePcbRenders [ERROR]: '{pcb_filename}' does not exist."))
    command_exec_status ["pcb_render"] = False
    return
  
  #---------------------------------------------------------------------------------------------#
  
  file_name = extract_pcb_file_name (pcb_filename) # Extract information from the input file
  file_name = file_name.replace (" ", "-") # If there are whitespace characters in the project name, replace them with a hyphen

  project_name = extract_project_name (file_name)
  info = extract_info_from_pcb (pcb_filename)
  
  print (f"generatePcbRenders [INFO]: Project name is '{color.magenta (project_name)}' and revision is {color.magenta ('R')}{color.magenta (info ['rev'])}.")

  #---------------------------------------------------------------------------------------------#

  # Check if the passed preset exists in the config file.
  if preset == None:
    print (color.yellow (f"generatePcbRenders [WARNING]: No render presets are specified in cli. All render presets available in the config file will be used."))

  else:
    # Check if the preset is valid.
    if preset not in current_config.get ("data", {}).get ("pcb_render", {}):
      print (color.red (f"generatePcbRenders [ERROR]: Invalid render preset '{preset}'. Please check the config file."))
      command_exec_status ["pcb_render"] = False
      return
    else:
      print (f"generatePcbRenders [INFO]: Using render preset '{color.magenta (preset)}' from the config file.")
    
  #---------------------------------------------------------------------------------------------#

  file_path = os.path.abspath (pcb_filename) # Get the absolute path of the file.

  # Get the directory path of the file and save it as the project directory.
  # All other export directories will be relative to the project directory.
  project_dir = os.path.dirname (file_path)
  
  # Read the output directory name from the config file.
  od_from_config = project_dir + "/" + current_config.get ("data", {}).get ("pcb_render", {}).get ("--output_dir", lambda: default_config ["data"]["pcb_render"]["--output_dir"])
  od_from_cli = output_dir  # The directory specified by the command line argument

  # Get the final directory path
  final_directory, filename_date = create_final_directory (od_from_config, od_from_cli, "Renders", info ["rev"], "generatePcbRenders")

  #---------------------------------------------------------------------------------------------#

  preset_list = []

  if preset == None:

    # Get the list of all render presets from the config file.
    for key, value in current_config.get ("data", {}).get ("pcb_render", {}).items():
      if not key.startswith ("--"):
        preset_list.append (key)
    
    # Print the list of render presets.
    print (f"generatePcbRenders [INFO]: Using all render presets: {color.blue (', '.join (preset_list))}")
  
  else:
    preset_list = [preset] # Create a list with the specified preset
  
  #---------------------------------------------------------------------------------------------#

  # Now render images using all of the presets.
  for preset in preset_list:
    full_command = []
    full_command.extend (pcb_render_export_command)

    seq_number = 1
    not_completed = True
    
    name_stub = current_config.get ("data", {}).get ("pcb_render", {}).get (preset, {}).get ("kie_name_stub", lambda: default_config ["data"]["pcb_render"][preset]["kie_name_stub"])
    kie_type = current_config.get ("data", {}).get ("pcb_render", {}).get (preset, {}).get ("kie_type", lambda: default_config ["data"]["pcb_render"][preset]["kie_type"])

    # Generate the file name.
    while not_completed:
      file_name = f"{final_directory}/{project_name}-R{info ['rev']}-{name_stub}-{filename_date}-{seq_number}.{kie_type}"

      if os.path.exists (file_name):
        seq_number += 1
        not_completed = True
      else:
        full_command.append ("--output")
        full_command.append (f'"{file_name}"') # Add the output file name with double quotes around it
        break
    
    # Get the argument list from the config file.
    arg_list = current_config.get ("data", {}).get ("pcb_render", {}).get (preset)

    # List of commands which should not be outputed as quoted strings but single comma string
    # Otherwise cannot handle negative values for angles
    list_of_comma_keys = ["--pan","--pivot","--rotate","--light-top","--light-bottom","--light-side","--light-camera"]
    
    # Add the remaining arguments.
    # Check if the argument list is not an empty dictionary.
    if arg_list:
      for key, value in arg_list.items():
        if key.startswith ("--"): # Only fetch the arguments that start with "--"
          if key == "--output_dir": # Skip the --output_dir argument, sice we already added it
            continue
          else:
            # Check if the value is empty
            if value == "": # Skip if the value is empty
              continue
            else:
              # Check if the vlaue is a JSON boolean
              if isinstance (value, bool):
                if value == True: # If the value is true, then append the key as an argument
                  full_command.append (key)
              else:
                # Check if the value is a string and not a numeral
                if isinstance (value, str) and not value.isdigit():
                    full_command.append (key)
                    if key in list_of_comma_keys:
                        full_command.append (f'\'{value}\'') # Add as a double-quoted string
                    else:
                        full_command.append (f'"{value}"') # Add as a double-quoted string
                elif isinstance (value, (int, float)):
                    full_command.append (key)
                    full_command.append (str (value))  # Append the numeric value as string
    
    # Finally add the input file
    full_command.append (f'"{pcb_filename}"')
    print ("generatePcbRenders [INFO]: Running command: ", color.blue (' '.join (full_command)))

    #---------------------------------------------------------------------------------------------#
  
    # Run the command
    try:
      full_command = ' '.join (full_command) # Convert the list to a string
      subprocess.run (full_command, check = True)
    
    except subprocess.CalledProcessError as e:
      print (color.red (f"generatePcbRenders [ERROR]: Error occurred: {e}"))
      command_exec_status ["pcb_render"] = False
      return

    print (color.green (f"generatePcbRenders [OK]: Render files using preset '{preset}' exported successfully."))
    print()
    command_exec_status ["pcb_render"] = True

#=============================================================================================#

def generateSchPdf (output_dir, sch_filename, pcb_filename = None, to_overwrite = True):
  global current_config  # Access the global config
  global default_config  # Access the global config

  # Get the KiCad CLI path.
  kicad_cli_path = f'{current_config.get ("kicad_cli_path", lambda: default_config ["kicad_cli_path"])}'

  # Common base command
  sch_pdf_export_command = [f'"{kicad_cli_path}"', "sch", "export", "pdf"]

  # Check if the input file exists
  if not check_file_exists (sch_filename):
    print (color.red (f"generateSchPdf [ERROR]: '{sch_filename}' does not exist."))
    command_exec_status ["sch_pdf"] = False
    return

  #---------------------------------------------------------------------------------------------#
  
  file_name = extract_pcb_file_name (sch_filename) # Extract information from the input file
  file_name = file_name.replace (" ", "-") # If there are whitespace characters in the project name, replace them with a hyphen

  project_name = extract_project_name (file_name)
  # Use pcb_filename rev info for dir structure when this is provided
  if pcb_filename != None:
    info = extract_info_from_pcb (pcb_filename) # Extract basic information from the input file (pcb)
  else:
    info = extract_info_from_pcb (sch_filename) # Extract basic information from the input file (sch)

  print (f"generateSchPdf [INFO]: Project name is '{color.magenta (project_name)}' and revision is {color.magenta ('R')}{color.magenta (info ['rev'])}.")

  #---------------------------------------------------------------------------------------------#

  file_path = os.path.abspath (sch_filename) # Get the absolute path of the file.

  # Get the directory path of the file and save it as the project directory.
  # All other export directories will be relative to the project directory.
  project_dir = os.path.dirname (file_path)
  
  # Read the output directory name from the config file.
  od_from_config = project_dir + "/" + current_config.get ("data", {}).get ("sch_pdf", {}).get ("--output_dir", lambda: default_config ["data"]["sch_pdf"]["--output_dir"])
  od_from_cli = output_dir  # The output directory specified by the command line argument

  # Get the final directory path.
  final_directory, filename_date = create_final_directory (od_from_config, od_from_cli, "SCH", info ["rev"], "generateSchPdf")
  
  #---------------------------------------------------------------------------------------------#
  
  # Get the argument list from the config file.
  arg_list = current_config.get ("data", {}).get ("sch_pdf", {})

  full_command = []
  full_command.extend (sch_pdf_export_command) # Add the base command

  seq_number = 1
  not_completed = True
  
  # Create the output file name.
  while not_completed:
    file_name = f"{final_directory}/{project_name}-R{info ['rev']}-SCH-{filename_date}-{seq_number}.pdf"

    if os.path.exists (file_name):
      seq_number += 1 # Increment the sequence number and try again
      not_completed = True
    else:
      full_command.append ("--output")
      full_command.append (f'"{file_name}"') # Add the output file name with double quotes around it
      break
  
  # Add the remaining arguments.
  # Check if the argument list is not an empty dictionary.
  if arg_list:
    for key, value in arg_list.items():
      if key.startswith ("--"): # Only fetch the arguments that start with "--"
        if key == "--output_dir": # Skip the --output_dir argument, sice we already added it
          continue
        else:
          # Check if the value is empty
          if value == "": # Skip if the value is empty
            continue
          else:
            # Check if the vlaue is a JSON boolean
            if isinstance (value, bool):
              if value == True: # If the value is true, then append the key as an argument
                full_command.append (key)
            else:
              # Check if the value is a string and not a numeral
              if isinstance (value, str) and not value.isdigit():
                  full_command.append (key)
                  full_command.append (f'"{value}"') # Add as a double-quoted string
              elif isinstance (value, (int, float)):
                  full_command.append (key)
                  full_command.append (str (value))  # Append the numeric value as string
  
  # Finally add the input file
  full_command.append (f'"{sch_filename}"')
  print ("generateSchPdf [INFO]: Running command: ", color.blue (' '.join (full_command)))

  #---------------------------------------------------------------------------------------------#
  
  # Run the command
  try:
    full_command = ' '.join (full_command) # Convert the list to a string
    subprocess.run (full_command, check = True)
  
  except subprocess.CalledProcessError as e:
    print (color.red (f"generateSchPdf [ERROR]: Error occurred: {e}"))
    print()
    command_exec_status ["sch_pdf"] = False
    return

  print (color.green ("generateSchPdf [OK]: Schematic PDF file exported successfully."))
  print()
  command_exec_status ["sch_pdf"] = True

#=============================================================================================#

def generate3D (output_dir, pcb_filename, type = "STEP", to_overwrite = True):
  # Get the KiCad CLI path.
  kicad_cli_path = f'{current_config.get ("kicad_cli_path", lambda: default_config ["kicad_cli_path"])}'

  # Common base command
  if type == "STEP" or type == "step":
    ddd_export_command = [f'"{kicad_cli_path}"', "pcb", "export", "step"]
    type = "STEP"
    extension = "step"
  elif type == "VRML" or type == "vrml":
    ddd_export_command = [f'"{kicad_cli_path}"', "pcb", "export", "vrml"]
    type = "VRML"
    extension = "wrl"

  if not check_file_exists (pcb_filename):
    print (color.red (f"generate3D [ERROR]: '{pcb_filename}' does not exist."))
    command_exec_status ["ddd" + "_" + type] = False
    return

  #---------------------------------------------------------------------------------------------#
  
  file_name = extract_pcb_file_name (pcb_filename) # Extract information from the input file
  file_name = file_name.replace (" ", "-") # If there are whitespace characters in the project name, replace them with a hyphen

  project_name = extract_project_name (file_name)
  info = extract_info_from_pcb (pcb_filename)
  
  print (f"generate3D [INFO]: Project name is '{color.magenta (project_name)}' and revision is {color.magenta ('R')}{color.magenta (info ['rev'])}.")

  #---------------------------------------------------------------------------------------------#

  file_path = os.path.abspath (pcb_filename) # Get the absolute path of the file.

  # Get the directory path of the file and save it as the project directory.
  # All other export directories will be relative to the project directory.
  project_dir = os.path.dirname (file_path)
  
  # Read the output directory name from the config file.
  od_from_config = project_dir + "/" + current_config.get ("data", {}).get ("ddd", {}).get (type, {}).get ("--output_dir", lambda: default_config ["data"]["ddd"][type]["--output_dir"])
  od_from_cli = output_dir  # The directory specified by the command line argument

  # Get the final directory path
  final_directory, filename_date = create_final_directory (od_from_config, od_from_cli, "3D", info ["rev"], "generate3D")

  #---------------------------------------------------------------------------------------------#
  
  full_command = []
  full_command.extend (ddd_export_command) # Add the base command
  
  seq_number = 1
  not_completed = True
  
  # Generate the file name.
  while not_completed:
    file_name = f"{final_directory}/{project_name}-R{info ['rev']}-{type}-{filename_date}-{seq_number}.{extension}"

    if os.path.exists (file_name):
      seq_number += 1
      not_completed = True
    else:
      full_command.append ("--output")
      full_command.append (f'"{file_name}"') # Add the output file name with double quotes around it
      break
  
  #---------------------------------------------------------------------------------------------#
  
  # Get the argument list from the config file.
  arg_list = current_config.get ("data", {}).get ("ddd", {}).get (type)

  # Add the remaining arguments.
  # Check if the argument list is not an empty dictionary.
  if arg_list:
    for key, value in arg_list.items():
      if key.startswith ("--"): # Only fetch the arguments that start with "--"
        if key == "--output_dir": # Skip the --output_dir argument, sice we already added it
          continue
        else:
          # Check if the value is empty
          if value == "": # Skip if the value is empty
            continue
          else:
            # Check if the vlaue is a JSON boolean
            if isinstance (value, bool):
              if value == True: # If the value is true, then append the key as an argument
                full_command.append (key)
            else:
              # Check if the value is a string and not a numeral
              if isinstance (value, str) and not value.isdigit():
                  full_command.append (key)
                  full_command.append (f'"{value}"') # Add as a double-quoted string
              elif isinstance (value, (int, float)):
                  full_command.append (key)
                  full_command.append (str (value))  # Append the numeric value as string
  
  # Finally add the input file
  full_command.append (f'"{pcb_filename}"')
  print ("generate3D [INFO]: Running command: ", color.blue (' '.join (full_command)))

  #---------------------------------------------------------------------------------------------#
  
  # Run the command
  try:
    full_command = ' '.join (full_command) # Convert the list to a string
    subprocess.run (full_command, check = True)
  
  except subprocess.CalledProcessError as e:
    print (color.red (f"generate3D [ERROR]: Error occurred: {e}"))
    print()
    command_exec_status ["ddd" + "_" + type] = False
    return

  print (color.green (f"generate3D [OK]: {type} file exported successfully."))
  print()
  command_exec_status ["ddd" + "_" + type] = True

#=============================================================================================#

def generateBomCsv (output_dir, sch_filename, to_overwrite = True):
  # Get the KiCad CLI path.
  kicad_cli_path = f'{current_config.get ("kicad_cli_path", lambda: default_config ["kicad_cli_path"])}'

  # Common base command
  bom_export_command = [f'"{kicad_cli_path}"', "sch", "export", "bom"]

  # Check if the input file exists
  if not check_file_exists (sch_filename):
    print (color.red (f"generateBomCsv [ERROR]: '{sch_filename}' does not exist."))
    command_exec_status ["bom_csv"] = False
    return False

  #---------------------------------------------------------------------------------------------#
  
  file_name = extract_pcb_file_name (sch_filename)
  file_name = file_name.replace (" ", "-") # If there are whitespace characters in the project name, replace them with a hyphen
  
  project_name = extract_project_name (file_name)
  info = extract_info_from_pcb (sch_filename) # Extract basic information from the input file (sch)
  
  print (f"generateBomCsv [INFO]: Project name is '{color.magenta (project_name)}' and revision is {color.magenta ('R')}{color.magenta (info ['rev'])}.")

  #---------------------------------------------------------------------------------------------#
  
  file_path = os.path.abspath (sch_filename) # Get the absolute path of the file.

  # Get the directory path of the file and save it as the project directory.
  # All other export directories will be relative to the project directory.
  project_dir = os.path.dirname (file_path)
  
  # Read the output directory name from the config file.
  od_from_config = project_dir + "/" + current_config.get ("data", {}).get ("bom", {}).get ("CSV").get ("--output_dir", lambda: default_config ["data"]["bom"]["CSV"]["--output_dir"])
  od_from_cli = output_dir  # The output directory specified by the command line argument

  # Get the final directory path.
  final_directory, filename_date = create_final_directory (od_from_config, od_from_cli, "BoM", info ["rev"], "generateBomCsv")
  
  #---------------------------------------------------------------------------------------------#
  
  full_command = []
  full_command.extend (bom_export_command) # Add the base command
  
  seq_number = 1
  not_completed = True
  
  # Create the output file name.
  while not_completed:
    file_name = f"{final_directory}/{project_name}-R{info ['rev']}-BoM-CSV-{filename_date}-{seq_number}.csv"

    if os.path.exists (file_name):
      seq_number += 1
      not_completed = True
    else:
      full_command.append ("--output")
      full_command.append (f'"{file_name}"') # Add the output file name with double quotes around it
      break

  #---------------------------------------------------------------------------------------------#

  # Get the argument list from the config file.
  arg_list = current_config.get ("data", {}).get ("bom", {}).get ("CSV")

  # Add the remaining arguments.
  # Check if the argument list is not an empty dictionary.
  if arg_list:
    for key, value in arg_list.items():
      if key.startswith ("--"): # Only fetch the arguments that start with "--"
        if key == "--output_dir": # Skip the --output_dir argument, sice we already added it
          continue
        else:
          # Check if the value is empty
          if value == "": # Skip if the value is empty
            continue
          else:
            # Check if the vlaue is a JSON boolean
            if isinstance (value, bool):
              if value == True: # If the value is true, then append the key as an argument
                full_command.append (key)
            else:
              # Check if the value is a string and not a numeral
              if isinstance (value, str) and not value.isdigit():
                full_command.append (key)
                full_command.append (f'"{value}"') # Add as a double-quoted string
              elif isinstance (value, (int, float)):
                full_command.append (key)
                full_command.append (str (value))  # Append the numeric value as string
  
  # Finally add the input file
  full_command.append (f'"{sch_filename}"')
  print ("generateBomCsv [INFO]: Running command: ", color.blue (' '.join (full_command)))

  #---------------------------------------------------------------------------------------------#
  
  # Run the command
  try:
    full_command = ' '.join (full_command) # Convert the list to a string
    subprocess.run (full_command, check = True)
  
  except subprocess.CalledProcessError as e:
    print (color.red (f"generateBomCsv [ERROR]: Error occurred: {e}"))
    print()
    command_exec_status ["bom_csv"] = False
    return False

  print (color.green ("generateBomCsv [OK]: BoM file exported successfully."))
  print()
  command_exec_status ["bom_csv"] = True
  return file_name # Return the file name of the generated BoM file

#=============================================================================================#
        
def generateBomXls (output_dir, csv_file, sch_filename, to_overwrite = True):
  # Check if the input schematic file exists
  if not check_file_exists (sch_filename):
    print (color.red (f"generateBomXls [ERROR]: '{sch_filename}' does not exist."))
    command_exec_status ["bom_csv"] = False
    return False
  
  # Check if the input CSV file exists
  if not check_file_exists (csv_file):
    print (color.red (f"generateBomXls [ERROR]: The supplied CSV file '{csv_file}' does not exist."))
    command_exec_status ["bom_xls"] = False
    return False

  print (f"generateBomXls [INFO]: CSV BoM file is '{color.magenta (csv_file)}'.")

  #---------------------------------------------------------------------------------------------#

  file_name = extract_pcb_file_name (sch_filename)
  file_name = file_name.replace (" ", "-") # If there are whitespace characters in the project name, replace them with a hyphen
  
  project_name = extract_project_name (file_name)
  info = extract_info_from_pcb (sch_filename) # Extract basic information from the input file (sch)
  
  print (f"generateBomXls [INFO]: Project name is '{color.magenta (project_name)}' and revision is {color.magenta ('R')}{color.magenta (info ['rev'])}.")

  #---------------------------------------------------------------------------------------------#

  file_path = os.path.abspath (sch_filename) # Get the absolute path of the file.

  # Get the directory path of the file and save it as the project directory.
  # All other export directories will be relative to the project directory.
  project_dir = os.path.dirname (file_path)
  
  # Read the output directory name from the config file.
  od_from_config = project_dir + "/" + current_config.get ("data", {}).get ("bom", {}).get ("XLS").get ("--output_dir", lambda: default_config ["data"]["bom"]["XLS"]["--output_dir"])
  od_from_cli = output_dir  # The output directory specified by the command line argument

  # Get the final directory path.
  final_directory, filename_date = create_final_directory (od_from_config, od_from_cli, "BoM", info ["rev"], "generateBomXls")

  #---------------------------------------------------------------------------------------------#

  seq_number = 1
  not_completed = True
  
  # Create the output file name.
  while not_completed:
    file_name = f"{final_directory}/{project_name}-R{info ['rev']}-BoM-XLS-{filename_date}-{seq_number}.xlsx"

    if os.path.exists (file_name):
      seq_number += 1
      not_completed = True
    else:
      break

  #---------------------------------------------------------------------------------------------#

  # Get the argument list from the config file.
  arg_list = current_config.get ("data", {}).get ("bom", {}).get ("XLS")
  
  # Set the default values for keys if ther are not listed in config file
  row_height = 50
  
  # Check if the argument list is not an empty dictionary.
  if arg_list:
    for key, value in arg_list.items():
      # Check for valid special keys defined for XLS BoM generation
      if key == "kie_row_height":
        if isinstance (value, str) and not value.isdigit():
          row_height = None                       # Any non integer value treated as None
        if isinstance (value, (int, float)):
          row_height = value
          if row_height == 0:                     # Value of 0 considered as None
            row_height = None
 
 # Create workbook and sheet
  wb = Workbook()
  ws = wb.active
  ws.title = "BoM"

  # Define styles
  header_font = Font (name = "Roboto Mono", bold = True)
  cell_font = Font (name = "Roboto Mono")
  alignment = Alignment (horizontal = "center", vertical = "center", wrap_text = True)
  header_fill = PatternFill (fill_type = "solid", fgColor = "CCE5FF")  # Light blue background

  # Custom widths mapping (column header => width in points)
  # Use a default setting if file not provided or is not correctly parsed
  def_custom_widths = {
    "#": 14,
    "Reference": 42,
    "Value": 34,
    "Name": 32,
    "Footprint": 62,
    "Qty": 16,
    "DNP": 16,
    "MPN": 31,
    "MFR": 28,
    "Alt MPN": 30,
  }
  # First try to load custom widths file from the current config file
  custom_widths = current_config.get ("data", {}).get ("bom", {}).get ("XLS").get ("kie_custom_widths")

  # Custom styles mapping (column header => style)
  # Use a default setting if file not provided or is not correctly parsed
  def_custom_styles = {
    "Link": "Hyperlink",
    "Datasheet": "Hyperlink",
    "URL": "Hyperlink"
  }
  # First try to load custom styles file from the current config file
  custom_styles = current_config.get ("data", {}).get ("bom", {}).get ("XLS").get ("kie_custom_styles")

  if len (custom_styles) == 0:
    # if this is not loaded as well then will use default custom styles dictionary
    custom_styles = def_custom_styles

  headers = []

  # Read CSV and write to worksheet
  with open (csv_file, newline = '', encoding = 'utf-8') as csvfile:
    reader = csv.reader (csvfile)
    for row_index, row in enumerate (reader, start = 1):
      # Store headers for width adjustment
      if row_index == 1:
        headers = row

      for col_index, value in enumerate (row, start = 1):
        cell = ws.cell (row = row_index, column = col_index, value = value)
        cell.font = header_font if row_index == 1 else cell_font
        cell.alignment = alignment
        style = custom_styles.get(headers[col_index-1], "None")
        # Special style Hyperlink also make the value with hyperlink contents)
        if style == "Hyperlink":
          cell.value = '=HYPERLINK("%s","%s")' % (value, value)
        if style != "None" and row_index > 1:
          cell.style = style
        if row_index == 1:
          cell.fill = header_fill
      
      # Set minimum row height from user parameter (None indicates auto height)
      ws.row_dimensions [row_index].height = row_height

  # Freeze the header row
  ws.freeze_panes = ws ["A2"]

  # Apply custom or fallback column widths
  for col_index, header in enumerate (headers, start = 1):
    col_letter = get_column_letter (col_index)
    width = custom_widths.get (header, 15)  # Default width if not specified
    ws.column_dimensions [col_letter].width = width

  # # Auto-adjust column widths
  # for col_index, column_cells in enumerate (ws.columns, start = 1):
  #   max_length = max (len (str (cell.value)) if cell.value else 0 for cell in column_cells)
  #   col_letter = get_column_letter (col_index)
  #   ws.column_dimensions [col_letter].width = max (10, min (max_length + 2, 40))  # Reasonable bounds

  # Save the styled XLSX
  wb.save (file_name)

  print (color.green ("generateBomXls [OK]: XLS BoM file exported successfully."))
  print()
  command_exec_status ["bom_xls"] = True

#=============================================================================================#

def generateSvg (output_dir, pcb_filename, to_overwrite = True):
  # Get the KiCad CLI path.
  kicad_cli_path = f'{current_config.get ("kicad_cli_path", lambda: default_config ["kicad_cli_path"])}'

  # Common base command
  pcb_svg_export_command = [f'"{kicad_cli_path}"', "pcb", "export", "svg"]

  # Check if the input file exists
  if not check_file_exists (pcb_filename):
    print (color.red (f"generateSvg [ERROR]: '{pcb_filename}' does not exist."))
    command_exec_status ["svg"] = False
    return

  #---------------------------------------------------------------------------------------------#
  
  file_name = extract_pcb_file_name (pcb_filename) # Extract information from the input file
  file_name = file_name.replace (" ", "-") # If there are whitespace characters in the project name, replace them with a hyphen

  project_name = extract_project_name (file_name)
  info = extract_info_from_pcb (pcb_filename) # Extract basic information from the input file

  print (f"generateSvg [INFO]: Project name is '{color.magenta (project_name)}' and revision is {color.magenta ('R')}{color.magenta (info ['rev'])}.")

  #---------------------------------------------------------------------------------------------#

  file_path = os.path.abspath (pcb_filename) # Get the absolute path of the file.

  # Get the directory path of the file and save it as the project directory.
  # All other export directories will be relative to the project directory.
  project_dir = os.path.dirname (file_path)
  
  # Read the output directory name from the config file.
  od_from_config = project_dir + "/" + current_config.get ("data", {}).get ("svg", {}).get ("--output_dir", lambda: default_config ["data"]["svg"]["--output_dir"])
  od_from_cli = output_dir  # The output directory specified by the command line argument

  # Get the final directory path.
  final_directory, filename_date = create_final_directory (od_from_config, od_from_cli, "SVG", info ["rev"], "generateSvg")
  
  #---------------------------------------------------------------------------------------------#

  # Delete the existing files in the output directory
  delete_files (final_directory, include_extensions = [".svg"])

  #---------------------------------------------------------------------------------------------#
  
  # Get the argument list from the config file.
  arg_list = current_config.get ("data", {}).get ("svg", {})

  # Check the number of technical layers to export. This is not the number of copper layers.
  layer_count = len (arg_list.get ("--layers", []))

  if layer_count <= 0:
    print (color.red (f"generateSvg [ERROR]: No layers specified for export."))
    command_exec_status ["svg"] = False
    return

  # Get the number of common layers to include in each of the PDF.
  # common_layer_count = len (arg_list.get ("kie_common_layers", []))

  base_command = []
  base_command.extend (pcb_svg_export_command) # Add the base command
  
  for i in range (layer_count):
    full_command = base_command [:]
    # Get the arguments.
    if arg_list: # Check if the argument list is not an empty dictionary.
      for key, value in arg_list.items():
        if key.startswith ("--"): # Only fetch the arguments that start with "--"
          if key == "--output_dir": # Skip the --output_dir argument, sice we already added it
            continue
          elif key == "--layers":
            layer_name = arg_list ["--layers"][i] # Get a layer name from the layer list
            layer_name = layer_name.replace (".", "_") # Replace dots with underscores
            layer_name = layer_name.replace (" ", "_") # Replace spaces with underscores

            full_command.append ("--output")
            full_command.append (f'"{final_directory}/{project_name}-R{info ["rev"]}-{layer_name}.svg"') # This is the ouput file name, and not a directory name

            layer_name = arg_list ["--layers"][i] # Get a layer name from the layer list
            layer_list = [f"{layer_name}"]  # Now create a list with the first item as the layer name
            common_layer_list = arg_list ["kie_common_layers"]  # Add the common layers
            layer_list.extend (common_layer_list) # Now combine the two lists
            layers_csv = ",".join (layer_list) # Convert the list to a comma-separated string
            full_command.append (key)
            full_command.append (f'"{layers_csv}"')
          else:
            # Check if the value is empty
            if value == "": # Skip if the value is empty
              continue
            else:
              # Check if the vlaue is a JSON boolean
              if isinstance (value, bool):
                if value == True: # If the value is true, then append the key as an argument
                  full_command.append (key)
              else:
                # Check if the value is a string and not a numeral
                if isinstance (value, str) and not value.isdigit():
                    full_command.append (key)
                    full_command.append (f'"{value}"') # Add as a double-quoted string
                elif isinstance (value, (int, float)):
                    full_command.append (key)
                    full_command.append (str (value))  # Append the numeric value as string

    full_command.append (f'"{pcb_filename}"')
    print ("generateSvg [INFO]: Running command: ", color.blue (' '.join (full_command)))

      # Run the command
    try:
      full_command = ' '.join (full_command) # Convert the list to a string
      subprocess.run (full_command, check = True)
      # print (color.green ("generateSvg [OK]: PCB SVG files exported successfully."))
    
    except subprocess.CalledProcessError as e:
      print (color.red (f"generateSvg [ERROR]: Error occurred: {e}"))
      command_exec_status ["svg"] = False
      continue

  #---------------------------------------------------------------------------------------------#

  print (color.green ("generateSvg [OK]: SVG files exported successfully."))
  command_exec_status ["svg"] = True

  #---------------------------------------------------------------------------------------------#
  
  seq_number = 1
  not_completed = True

  files_to_include = [".svg"]
  
  # Sequentially name and create the zip files.
  while not_completed:
    zip_file_name = f"{project_name}-R{info ['rev']}-PCB-SVG-{filename_date}-{seq_number}.zip"

    if os.path.exists (f"{final_directory}/{zip_file_name}"):
      seq_number += 1
    else:
      zip_cnt = zip_all_files_ext (final_directory, zip_file_name, files_to_include)
      print (f"generateSvg [OK]: ZIP file '{color.magenta (zip_file_name)}' created successfully with {zip_cnt} files.")
      print()
      not_completed = False

#=============================================================================================#

def create_final_directory (dir_from_config, dir_from_cli, target_dir_name, rev, func_name, to_overwrite = True):
  # This will be the root directory for the output files.
  # Extra directories will be created based on the revision, date and sequence number.
  target_dir = None

  # The command line directory has precedence over the  configured directory.
  if dir_from_cli == "":
    print (f"{func_name} [INFO]: CLI directory '{color.magenta (dir_from_config)}' is empty. Using the configured directory.")
    target_dir = dir_from_config # If it's empty, use the configured directory
  else:
    print (f"{func_name} [INFO]: CLI directory '{color.magenta (dir_from_config)}' is not empty. Using the CLI directory.")
    target_dir = dir_from_cli # Otherwise, use the command line argument

  if not os.path.exists (target_dir): # Check if the target directory exists
    print (f"{func_name} [INFO]: Output directory '{color.magenta (target_dir)}' does not exist. Creating it now.")
    os.makedirs (target_dir)
  else:
    print (f"{func_name} [INFO]: Output directory '{color.magenta (target_dir)}' already exists.")

  #---------------------------------------------------------------------------------------------#

  # Create one more directory based on the revision number.
  rev_directory = f"{target_dir}/R{rev}"

  # Check if the revision directory exists, and create if not.
  if not os.path.exists (rev_directory):
    print (f"{func_name} [INFO]: Revision directory '{color.magenta (rev_directory)}' does not exist. Creating it now.")
    os.makedirs (rev_directory)
  
  #---------------------------------------------------------------------------------------------#
  
  not_completed = True
  seq_number = 0
  
  # Now we have to make the date-specific and output-specific directory.
  # This will be the final directory for the output files.
  while not_completed:
    today_date = datetime.now()
    # formatted_date = today_date.strftime ("%d-%m-%Y")
    formatted_date = today_date.strftime ("%Y-%m-%d")
    filename_date = today_date.strftime ("%d%m%Y")
    seq_number += 1
    # date_directory = f"{rev_directory}/[{seq_number}] {formatted_date}"
    date_directory = f"{rev_directory}/{formatted_date}"
    final_directory = f"{date_directory}/{target_dir_name}"

    if not os.path.exists (final_directory):
      print (f"{func_name} [INFO]: Target directory '{color.magenta (final_directory)}' does not exist. Creating it now.")
      os.makedirs (final_directory)
      not_completed = False
    else:
      if to_overwrite:
        print (f"{func_name} [INFO]: Target directory '{color.magenta (final_directory)}' already exists. Files may be overwritten.")
        not_completed = False
      else:
        print (f"{func_name} [INFO]: Target directory '{color.magenta (final_directory)}' already exists. Creating another one.")
        not_completed = True
  
  return final_directory, filename_date

#=============================================================================================#

def load_dict_from_text_file(filepath, delimiter=':'):
    """
    Loads a dictionary from a text file with key:value pairs.
    In any case of error file not found or file not properly parsed (dictionary loaded has less than 2 items) an empty dic will be returned
    
    Args:
        filepath (str): Full path to the text file from which to load the key:value pairs to dictionary
        delimiter (char): optional delimiter, default is ':'
    """
    data = {}
    
    try:
        with open(filepath, 'r') as f:
            for line in f:
                line = line.strip()
                if line and delimiter in line:
                    key, value = line.split(delimiter, 1)
                    data[key.strip()] = value.strip()
        if len (data) > 1:  # Expected min 2 valid parsed entries in file
            return data
        return {}
    except: # Any exception file not found or other
        return {}

#=============================================================================================#

def zip_all_files_ext (source_folder, zip_file_name = None, extensions = None):
    """
    Compresses files from a folder into a ZIP file, including only files with specified extensions.

    Args:
        source_folder (str): Path to the folder containing files.
        extensions (list of str, optional): List of file extensions to include (e.g., ['.txt', '.jpg']) (None means Zip all).
        zip_file_name (str, optional): Name of the ZIP file. If None, will use 'archive.zip'.

    Return:
        zip_cnt (int): Total number of files zipped to zip file
    """
    if extensions is None:
        extensions = []  # Include all files if no extensions are specified
    
    if zip_file_name is None:
        zip_file_name = 'archive.zip'  # Default ZIP file name
    
    zip_file_path = os.path.join (source_folder, zip_file_name)
    zip_cnt = 0
    
    with zipfile.ZipFile (zip_file_path, 'w') as zipf:
        for foldername, subfolders, filenames in os.walk (source_folder):
            for filename in filenames:
                file_path = os.path.join (foldername, filename)
                # Check if the file has one of the specified extensions
                if not extensions or any (filename.endswith (ext) for ext in extensions):
                    # Exclude the ZIP file itself from being added
                    if os.path.abspath (file_path) != os.path.abspath (zip_file_path):
                        zipf.write (file_path, arcname = os.path.relpath (file_path, source_folder))
                        zip_cnt = zip_cnt + 1
    
    # print(f"ZIP file created: {zip_file_name}")

    return zip_cnt

#=============================================================================================#

def zip_all_files_list (source_folder, zip_file_path, file_list = None):
  """
  Compresses files from a folder into a ZIP file, including only files within provided list.

  Args:
      source_folder (str): Path to the folder containing files.
      file_list (list of str, optional): List of file name (in source dir) to include (None means Zip All).
      zip_file_path (str): Path where the ZIP file will be saved.

  Return:
      zip_cnt (int): Total number of files zipped to zip file
  """
  
  zip_cnt = 0  
  with zipfile.ZipFile (zip_file_path, 'w') as zipf:
    for foldername, subfolders, filenames in os.walk (source_folder):
      for filename in filenames:
        if foldername == source_folder:
          # Include only files from file list in source_folder (or all if list is none, except for .zip files)
          if (file_list is None and not filename.endswith('.zip')) or f'{filename}' in file_list:
            file_path = os.path.join (foldername, filename)
            zipf.write (file_path, arcname = os.path.relpath (file_path, source_folder))
            zip_cnt = zip_cnt + 1

  # print(f"ZIP file created: {os.path.basename (zip_file_path)}")

  return zip_cnt

#=============================================================================================#

def delete_non_zip_files (directory):
  """
  Deletes all files in the specified directory except ZIP files.

  Args:
    directory (str): Path to the directory where the cleanup will occur.
  """
  for filename in os.listdir (directory):
    file_path = os.path.join (directory, filename)
    if os.path.isfile (file_path) and not filename.endswith ('.zip'):
      os.remove (file_path)
      # print(f"Deleted: {filename}")

#=============================================================================================#

def delete_files_with_extensions (directory, extensions = None):
    """
    Deletes files in the specified directory with the specified extensions.

    Args:
        directory (str): Path to the directory where the cleanup will occur.
        extensions (str or list of str, optional): Comma-separated string or list of file extensions to delete.
    """
    if extensions is None:
        # print ("No extensions provided. No files will be deleted.")
        return
    
    if isinstance (extensions, str):
        extensions = extensions.split (',')

    # Ensure all extensions are in the form of '.ext'
    extensions = [f".{ext.strip()}" for ext in extensions]

    for filename in os.listdir (directory):
        file_path = os.path.join (directory, filename)
        if os.path.isfile (file_path):
            # Check if file extension matches one of the provided extensions
            if any (filename.endswith (ext) for ext in extensions):
                os.remove (file_path)
                # print(f"Deleted: {filename}")

#=============================================================================================#

def delete_files (directory, include_extensions = None, exclude_extensions = None):
    """
    Deletes files in the specified directory based on the inclusion and exclusion lists of file extensions.

    Args:
        directory (str): Path to the directory where the cleanup will occur.
        include_extensions (list of str, optional): List of file extensions to include for deletion, e.g., ['.txt', '.jpg'].
        exclude_extensions (list of str, optional): List of file extensions to exclude from deletion, e.g., ['.zip'].
    """
    # Ensure inclusion and exclusion lists are properly formatted
    if include_extensions is None:
        include_extensions = []
    # Ensure that include_extensions have leading dots and are unique
    include_extensions = [ext.strip().lower() for ext in include_extensions if ext.startswith('.')]
    
    if exclude_extensions is None:
        exclude_extensions = []
    # Ensure that exclude_extensions have leading dots and are unique
    exclude_extensions = [ext.strip().lower() for ext in exclude_extensions if ext.startswith('.')]

    for filename in os.listdir (directory):
        file_path = os.path.join (directory, filename)
        if os.path.isfile (file_path):
            # Get the file extension
            file_ext = os.path.splitext (filename) [1].lower()
            # Check if file extension is in the inclusion list and not in the exclusion list
            if (not include_extensions or file_ext in include_extensions) and (file_ext not in exclude_extensions):
                os.remove(file_path)
                # print(f"Deleted: {filename}")

#=============================================================================================#

def delete_directory (directory):
    """
    Deletes a complete directory tree with all its files and subdirs wheather empty or not
    This cannot be undone and does not save contents to recycle bin

    Args:
        directory (str): Path to the directory to be deleted
    """

    try:
        if os.path.isdir(directory):
            shutil.rmtree(directory)
        else:
            print(f"Directory '{directory}' does not exist, nothing to delete.")
    except OSError as e:
        print (color.red (f"delete_directory [ERROR]: Error while deleting directory: {e.filename} - {e.strerror}."))

#=============================================================================================#

def rename_files (directory, prefix, revision = "", extensions = None):
  """
  Rename files in the specified directory. Only files starting with the prefix will be renamed.
  You can also specify the revision number to be used when the file is renamed. The renaming will only
  be applied to the files with the listed extensions. If no extensions are given,
  all files will be renamed.

  Args:
      directory (str): The directory of the files to rename.
      prefix (str): The prefix of the files to be renamed. Only these files will be renamed.
      revision (str, optional): The revision tag. For example, "0.1". Defaults to "".
      extensions (list of str, optional): The list of extensions. Only these files will be renamed. Defaults to None.
  """
  
  if extensions is None:
    extensions = [] # All file types will be considered if no extension filter is specified

  for filename in os.listdir (directory):
    # Check if the filename starts with the prefix and ends with a valid extension (or any extension if none specified)
    if filename.startswith (prefix) and (not extensions or any (filename.endswith (ext) for ext in extensions)):
      # Construct the new filename with the revision tag
      base_name = filename [len (prefix):]  # Remove the prefix part
      new_filename = f"{prefix}-R{revision}{base_name}"
      
      # Full paths for renaming
      old_file_path = os.path.join (directory, filename)
      new_file_path = os.path.join (directory, new_filename)
      
      # Rename the file
      os.rename (old_file_path, new_file_path)
      # print(f"Renamed: {filename} -> {new_filename}")

#=============================================================================================#

def check_file_exists (file_name):
  """
  Checks if the input PCB file exists.
  Args:
    input_file_name (str): The path to the PCB file.
  Returns:
    bool: True if the file exists, False otherwise.
  """
  return os.path.exists (file_name)

#=============================================================================================#

def extract_project_name (file_name):
  """
  Extracts the project name from a given PCB file name by removing the extension.
  Args:
    input_file_name (str): The PCB file name.
  Returns:
    str: The project name without the file extension.
  """
  file_name = extract_pcb_file_name (file_name)
  project_name = os.path.splitext (file_name) [0]
  # print ("Project name is ", project_name)

  return project_name

#=============================================================================================#

def extract_pcb_file_name (file_name):
  """
  Extracts the PCB file name from a given path.
  Args:
    input_file_name (str): The path to the PCB file.
  Returns:
    str: The PCB file name without any directory path.
  """
  return os.path.basename (file_name)

#=============================================================================================#

def extract_info_from_pcb (pcb_file_path):
  """
  Extracts specific information from a KiCad PCB file.
  Args:
    pcb_file_path (str): Path to the KiCad PCB file.
  Returns:
    dict: A dictionary containing the extracted information.
  """
  info = {}
  
  try:
    with open (pcb_file_path, 'r', encoding = "utf-8") as file:
      content = file.read()
    
    # Regular expressions to extract information
    title_match = re.search (r'\(title "([^"]+)"\)', content)
    date_match = re.search (r'\(date "([^"]+)"\)', content)
    rev_match = re.search (r'\(rev "([^"]+)"\)', content)
    company_match = re.search (r'\(company "([^"]+)"\)', content)
    comment1_match = re.search (r'\(comment 1 "([^"]+)"\)', content)
    comment2_match = re.search (r'\(comment 2 "([^"]+)"\)', content)
    
    # Store matches in the dictionary
    if title_match:
      info ['title'] = title_match.group (1)
    if date_match:
      info ['date'] = date_match.group (1)
    if rev_match:
      info ['rev'] = rev_match.group (1)
    if company_match:
      info ['company'] = company_match.group (1)
    if comment1_match:
      info ['comment1'] = comment1_match.group (1)
    if comment2_match:
      info ['comment2'] = comment2_match.group (1)
      
  except FileNotFoundError:
    print (f"Error: The file '{pcb_file_path}' does not exist.")
  except Exception as e:
    print (f"Error occurred: {e}")
  
  return info

#=============================================================================================#

def validate_command_list (cli_string):
  """
  Validates the command list from CLI against a list of valid commands.

  Args:
    `cli_string` (`str`): A list of commands, for example "gerbers, drills, sch_pdf".
    or "gerbers, drills, sch_pdf, [ddd, STEP]".

  Returns:
    `False`: If the command list is invalid.
    `validated_list` ([]) : A list of valid commands.
  """

  # Top level commands.
  valid_commands_json = json.dumps ({
    "pcb_drc": ["report", "json"],
    "sch_erc": ["report", "json"],
    "gerbers": [],
    "drills": [],
    "sch_pdf": [],
    "bom": ["CSV", "XLS", "HTML"],
    "pcb_pdf": [],
    "pcb_fab_pdf": [],
    "positions": [],
    "svg": [],
    "ddd": ["STEP", "VRML"],
    "pcb_render": [],
    "source_zip": []
  })

  # Some commands like the "pcb_render" can have custom named presets.
  # We don't need to check them.
  preset_commands_json = json.dumps ({
    "pcb_render": []
  })
  
  #---------------------------------------------------------------------------------------------#

  def quote_bare_words (s):
    def replacer (match):
        word = match.group (0)
        return f'"{word}"'
    pattern = r'\b(?!True|False|None)[\w-]+\b'
    return re.sub (pattern, replacer, s)

  #---------------------------------------------------------------------------------------------#

  # Load JSON dict
  valid_commands_dict = json.loads (valid_commands_json)

  # Parse CLI string to Python list
  try:
    safe_str = quote_bare_words (cli_string)
    parsed_cli = ast.literal_eval (f'[{safe_str}]')
  except Exception as e:
    print (color.red (f"validate_command_list [ERROR]: Failed to parse CLI input: {e}"))
    return False

  # Validate
  validated_list = []

  for item in parsed_cli:
    # Standalone commands. 
    if isinstance (item, str):
      if item not in valid_commands_dict: # Check if the command is a valid top-level command
        print (color.red (f"validate_command_list [ERROR]: Invalid standalone command '{item}'"))
        return False
      
      validated_list.append (item)

    # Commands with subcommands.
    elif isinstance (item, list) and len (item) == 2:
      main, sub = item
      if main not in valid_commands_dict: # Check if the command is a valid top-level command
        print (color.red (f"validate_command_list [ERROR]: Invalid main command '{main}'"))
        return False
      
      if (main not in preset_commands_json) and (sub not in valid_commands_dict [main]):
        print (color.red (f"validate_command_list [ERROR]: Invalid subcommand '{sub}' for main command '{main}'"))
        return False
      
      validated_list.append ([main, sub])
    else:
      print (color.red (f"validate_command_list [ERROR]: Unrecognized command format '{item}'"))
    
  return validated_list

#=============================================================================================#

def load_config (config_file = None):
  """
  Loads the JSON configuration file. If no file is provided, it uses the default configuration.

  Args:
    config_file (str, optional): Path to the configuration file. Can be relative or absolute. Defaults to None.
  """
  
  global current_config  # Declare the global variable here
  global default_config  # Declare the global variable here

  # Load the default configuration.
  # This is required to load values that are missing in the user-provided configuration.
  print (f"load_config [INFO]: Loading default configuration.")
  default_config = json.loads (DEFAULT_CONFIG_JSON)
  default_config = to_lazy_dict (default_config)

  #---------------------------------------------------------------------------------------------#
  
  if config_file is not None:
    # Load the configuration from the specified file.
    if os.path.exists (config_file):
      print (f"load_config [INFO]: Loading configuration from '{color.magenta (config_file)}'.")
      with open (config_file, 'r', encoding = "utf-8") as file:
          current_config = json.load (file)
          current_config = to_lazy_dict (current_config)
          # Compare the configuration JSON versions using semver library.
          if "version" in current_config:
            if semver.compare (normalize_version (MIN_CONFIG_JSON_VERSION), normalize_version (current_config ["version"])) == 0:
              print (f"load_config [INFO]: The configuration file version '{color.magenta (current_config ['version'])}' is the same as the minimum supported version '{MIN_CONFIG_JSON_VERSION}'.")
              return True
            elif semver.compare (normalize_version (MIN_CONFIG_JSON_VERSION), normalize_version (current_config ["version"])) < 0:
              print (color.yellow (f"load_config [WARNING]: The configuration file version '{current_config ['version']}' is newer than the minimum supported version '{MIN_CONFIG_JSON_VERSION}'."))
              return True
            elif semver.compare (normalize_version (MIN_CONFIG_JSON_VERSION), normalize_version (current_config ["version"])) > 0:
              print (color.red (f"load_config [ERROR]: The configuration file version '{current_config ['version']}' is older than the minimum supported version '{MIN_CONFIG_JSON_VERSION}'."))
              return False
    else:
      print (color.red (f"load_config [ERROR]: The provided configuration file '{config_file}' does not exist."))
      print (color.yellow (f"load_config [WARNING]: Default values will be used. Continue? [Y/N]"))
      user_input = input ("").strip().lower()
      if user_input != "y": return False
      current_config = default_config
      return True

  else:
    # Use the default configuration.
    print (f"load_config [INFO]: Using default configuration.")
    current_config = default_config
    return True

#=============================================================================================#

def normalize_version (version: str) -> str:
  """
  Normalize a version string to full semantic version format (MAJOR.MINOR.PATCH).
  Examples:
    "1"     -> "1.0.0"
    "1.6"   -> "1.6.0"
    "1.6.2" -> "1.6.2"
  """
  parts = version.strip().split ('.')
  while len (parts) < 3:
    parts.append ('0')
  return '.'.join (parts [:3])  # Ensure no more than 3 parts

#=============================================================================================#

def run (config_file, command_list = None, clean = False):
  """
  Fetches command list and configuration from the JSON file and runs the commands.

  Args:
    config_file (str): Path to the configuration file. Can be relative or absolute.
    command_list (str): string of the commands to be executed as part of this run job request
    clean (bool): True indicates to do a complete directory delete of the current rev export tree before creating al new export files
  """

  print (f"run [INFO]: Running KiExport with configuration file '{color.magenta (config_file)}'.")

  # Check if the configuration is loaded successfully.
  if (load_config (config_file) is not True):
    print (color.red ("run [ERROR]: Failed to load the configuration file."))
    return

  #---------------------------------------------------------------------------------------------#

  # List of top-level commands.
  valid_commands = ["pcb_drc", "gerbers", "drills", "sch_erc", "sch_pdf", "bom", "pcb_pdf", "pcb_fab_pdf", "positions", "ddd", "svg", "pcb_render", "source_zip"]

  #---------------------------------------------------------------------------------------------#

  # Get the command list from the config file.
  config_cmd_list = current_config.get ("commands", [])
  config_cmd_strings = []
  config_cmd_lists = []

  # Check the commands in the configuration file.
  if not config_cmd_list:
    print (color.red ("run [ERROR]: No list of commands specified in the configuration file."))
    return
  else:
    cmd_count = 0

    # Check if the commands are valid.
    for cmd in config_cmd_list:
      # If cmd is a string (eg. "drills"), validate directly.
      if isinstance (cmd, str) and cmd in valid_commands:
        config_cmd_strings.append (cmd)
        cmd_count += 1

      # If cmd is a list (eg. "["ddd", "STEP"]"), validate the first item as a command.
      elif isinstance (cmd, list) and cmd [0] in valid_commands:
        config_cmd_lists.append (cmd)
        cmd_count += 1

    if cmd_count == 0:
      print (color.red ("run [ERROR]: No valid commands specified in the configuration file."))
      return
    else:
      # Print the validated commands.
      print (f"run [INFO]: Found the following commands in the configuration file: {color.green (config_cmd_list)}, {color.green (cmd_count)}.")
  
  #---------------------------------------------------------------------------------------------#

  if command_list is not None:
    # Get the command list from the cli.
    cli_cmd_list = validate_command_list (cli_string = command_list)

    if cli_cmd_list is False:
      print (color.red ("run [ERROR]: Invalid command list provided."))
      return

    cli_cmd_strings = []
    cli_cmd_lists = []

    # Now check if the commands passed through the CLI are valid.
    if not cli_cmd_list:
      print (color.reset ("run [INFO]: No command subset provided. Running all commands from the configuration file."))
      # return
    else:
      cmd_count = 0

      # Check if the commands are valid.
      for cmd in cli_cmd_list:
        # If cmd is a string (eg. "drills"), validate directly.
        if isinstance (cmd, str) and cmd in valid_commands:
          cli_cmd_strings.append (cmd)
          cmd_count += 1

        # If cmd is a list (eg. "["ddd", "STEP"]"), validate the first item as a command.
        elif isinstance (cmd, list) and cmd [0] in valid_commands:
          cli_cmd_lists.append (cmd)
          cmd_count += 1

      if cmd_count == 0:
        print (color.reset ("run [INFO]: No command subset provided. Running all commands from the configuration file."))
        # return
      else:
        # Print the validated commands.
        print (f"run [INFO]: Found the following commands in the cli list: {color.green (cli_cmd_list)}, {color.green (cmd_count)}.")

  #---------------------------------------------------------------------------------------------#

  # Choose the list of commands.
  if command_list is not None:
    cmd_strings = cli_cmd_strings
    cmd_lists = cli_cmd_lists
  else:
    cmd_strings = config_cmd_strings
    cmd_lists = config_cmd_lists

  # Find the absolute path to the config file directory.
  config_path = os.path.abspath (config_file)
  project_dir = os.path.dirname (config_path)
  project_name = current_config.get ("project_name")

  # Special handling to automatically detect project name from a single kicad_pcb/kicad_sch pair files found in current config directory
  if project_name == None or project_name == "" or project_name == "None" or project_name == "Auto":
    # Look for kicad_pcb and kicad_sch in currect config dir
    entries = os.listdir(project_dir)

    pcb_filename = None
    sch_filename = None
    for entry in entries:
        full_path = os.path.join(project_dir, entry)
        if os.path.isfile(full_path):
            file_name = os.path.basename(full_path)
            root, ext = os.path.splitext(file_name)
            if ext == ".kicad_pcb":
                if pcb_filename != None:
                    # More than one kicad_pcb file exist in directory so cannot determine what is the project name
                    print (color.red ("run [ERROR]: Cannot determine the project name to use in project directory {project_dir}."))
                    return
                pcb_filename = root
            if ext == ".kicad_sch":
                if sch_filename != None:
                    # More than one kicad_sch file exist in directory so cannot determine what is the project name
                    print (color.red ("run [ERROR]: Cannot determine the project name to use in project directory {project_dir}."))
                    return
                sch_filename = root
                
    if pcb_filename == None and sch_filename == None:
        # No kicad_sch or kicad_pcb file exist in directory so cannot determine what is the project name
        print (color.red ("run [ERROR]: Cannot determine the project name to use in project directory {project_dir}."))
        return
    if pcb_filename == None:
        # Only kicad_sch file found in directory use it as the project name
        project_name = sch_filename
    elif sch_filename == None:
        # Only kicad_pcb file found in directory use it as the project name
        project_name = pcb_filename
    else:
        # Both kicad_pcb and kicad_sch files found in directory use either ONLY if they have the same root name
        if pcb_filename != sch_filename:
            print (color.red ("run [ERROR]: Cannot determine the project name to use in project directory {project_dir}."))
            return
        project_name = pcb_filename

  pcb_filename = project_name + ".kicad_pcb"
  sch_filename = project_name + ".kicad_sch"

  pcb_file_path = project_dir + "\\" + pcb_filename
  sch_file_path = project_dir + "\\" + sch_filename

  # Print the file names.
  print (f"run [INFO]: Project Directory: {color.magenta (project_dir)}")
  print (f"run [INFO]: PCB File: {color.magenta (pcb_filename)}")
  print (f"run [INFO]: Schematic File: {color.magenta (sch_filename)}")
  print()

  # Check if the file exists.
  if not os.path.isfile (pcb_file_path):
    print (color.yellow (f"run [WARNING]: The PCB file '{pcb_file_path}' does not exist."))
    print (color.yellow ("run [WARNING]: Commands that require the PCB file won't be executed."))
    info_pcb = None
  else:
    info_pcb = extract_info_from_pcb (pcb_filename) # Extract basic information from the input file
  if not os.path.isfile (sch_file_path):
    print (color.yellow (f"run [WARNING]: The Schematic file '{sch_file_path}' does not exist."))
    print (color.yellow ("run [WARNING]: Commands that require the Schematic file won't be executed."))
    info_sch = None
  else:
    info_sch = extract_info_from_pcb (sch_filename) # Extract basic information from the input file

  #---------------------------------------------------------------------------------------------#
  # Check SCH and PCB (if both provided) have same version info (as recommended) 
  if info_sch != None and info_pcb != None:
    if info_sch ["rev"] != info_pcb ["rev"]:
      print (color.yellow (f"run [WARNING]: The Schematic file and PCB file have different revisions, to avoid confusion it is recommended to keep unique project version in both. Do you want to continue? [Y/N]"))
      user_input = input ("").strip().lower()
      if user_input == "y":
        print (color.yellow ("run [WARNING]: Process will continue with different PCB and Schematic revision."))
      else:
        print (color.yellow ("run [WARNING]: Process will exit."))
        exit (1)  # Exit the script with an error code if the user does not want to continue

  # If requested clean the pcb revision from all previous runs
  if clean == True:
    if info_pcb != None:
      # Clean will only be preformed if PCB exists since considered the major release source
      # Clean will only be performed if atleast one of the following commands exists in cmd list: gerbers, sch_pdf, pcb_pdf, pcb_fab_pdf, bom
      # Construct the current rev output directory as would be constructed by each command run
      # The command line directory has precedence over the configured directory.
      # Taking the output directory from first command in command list - assume all same directory output !!
      output_dir = None
      for cmd in cmd_strings:
        if cmd == "gerbers":
          output_dir = current_config.get ("data", {}).get ("gerbers", {}).get ("--output_dir", lambda: default_config ["data"]["gerbers"]["--output_dir"])

        elif cmd == "sch_pdf":
          output_dir = current_config.get ("data", {}).get ("sch_pdf", {}).get ("--output_dir", lambda: default_config ["data"]["sch_pdf"]["--output_dir"])

        elif cmd == "bom":
          # Default is CSV
          output_dir = current_config.get ("data", {}).get ("bom", {}).get ("CSV", {}).get ("--output_dir", lambda: default_config ["data"]["bom"]["CSV"]["--output_dir"])

        elif cmd == "pcb_pdf":
          output_dir = current_config.get ("data", {}).get ("pcb_pdf", {}).get ("--output_dir", lambda: default_config ["data"]["pcb_pdf"]["--output_dir"])

        elif cmd == "pcb_fab_pdf":
          output_dir = current_config.get ("data", {}).get ("pcb_fab_pdf", {}).get ("--output_dir", lambda: default_config ["data"]["pcb_fab_pdf"]["--output_dir"])

      if output_dir == None:
        for cmd in cmd_lists:
          if cmd[0] == "gerbers":
            output_dir = current_config.get ("data", {}).get ("gerbers", {}).get ("--output_dir", lambda: default_config ["data"]["gerbers"]["--output_dir"])

          elif cmd[0] == "sch_pdf":
            output_dir = current_config.get ("data", {}).get ("sch_pdf", {}).get ("--output_dir", lambda: default_config ["data"]["sch_pdf"]["--output_dir"])

          elif cmd[0] == "bom":
            if cmd [1] == "CSV":
              output_dir = current_config.get ("data", {}).get ("bom", {}).get ("CSV", {}).get ("--output_dir", lambda: default_config ["data"]["bom"]["CSV"]["--output_dir"])
            elif cmd [1] == "XLS":
              output_dir = current_config.get ("data", {}).get ("bom", {}).get ("XLS", {}).get ("--output_dir", lambda: default_config ["data"]["bom"]["XLS"]["--output_dir"])

          elif cmd[0] == "pcb_pdf":
            output_dir = current_config.get ("data", {}).get ("pcb_pdf", {}).get ("--output_dir", lambda: default_config ["data"]["pcb_pdf"]["--output_dir"])

          elif cmd[0] == "pcb_fab_pdf":
            output_dir = current_config.get ("data", {}).get ("pcb_fab_pdf", {}).get ("--output_dir", lambda: default_config ["data"]["pcb_fab_pdf"]["--output_dir"])

      if output_dir != None:
        target_dir = project_dir + "/" + output_dir
        rev = info_pcb ["rev"]
        rev_directory = f"{target_dir}/R{rev}"
        delete_directory (rev_directory)
    else:
      print (f"run [INFO]: PCB file not found, clean request is ignored")
  #---------------------------------------------------------------------------------------------#

  csv_file_created = False
  csv_filename = None
  src_output_dir = None

  # Process the commands without any arguments or modifiers. eg. "gerbers", "drills", "sch_pdf", etc.
  for cmd in cmd_strings:
    if cmd == "pcb_drc":
      output_dir = current_config.get ("data", {}).get ("pcb_drc", {}).get ("--output_dir", lambda: default_config ["data"]["pcb_drc"]["--output_dir"])
      output_dir = project_dir + "\\" + output_dir # Output directory is relative to the project directory
      runDRC (output_dir = output_dir, pcb_filename = pcb_file_path, type = "report") # Default is "report"

    elif cmd == "gerbers":
      output_dir = current_config.get ("data", {}).get ("gerbers", {}).get ("--output_dir", lambda: default_config ["data"]["gerbers"]["--output_dir"])
      output_dir = project_dir + "\\" + output_dir # Output directory is relative to the project directory
      generateGerbers (output_dir = output_dir, pcb_filename = pcb_file_path)

    elif cmd == "drills":
      output_dir = current_config.get ("data", {}).get ("drills", {}).get ("--output_dir", lambda: default_config ["data"]["drills"]["--output_dir"])
      output_dir = project_dir + "\\" + output_dir # Output directory is relative to the project directory
      generateDrills (output_dir = output_dir, pcb_filename = pcb_file_path)

    elif cmd == "sch_erc":
      output_dir = current_config.get ("data", {}).get ("sch_erc", {}).get ("--output_dir", lambda: default_config ["data"]["sch_erc"]["--output_dir"])
      output_dir = project_dir + "\\" + output_dir # Output directory is relative to the project directory
      runERC (output_dir = output_dir, sch_filename = sch_file_path, type = "report") # Default is "report"

    elif cmd == "sch_pdf":
      output_dir = current_config.get ("data", {}).get ("sch_pdf", {}).get ("--output_dir", lambda: default_config ["data"]["sch_pdf"]["--output_dir"])
      output_dir = project_dir + "\\" + output_dir # Output directory is relative to the project directory
      generateSchPdf (output_dir = output_dir, sch_filename = sch_file_path)

    elif cmd == "bom":
      # Default is CSV
      output_dir = current_config.get ("data", {}).get ("bom", {}).get ("CSV", {}).get ("--output_dir", lambda: default_config ["data"]["bom"]["CSV"]["--output_dir"])
      output_dir = project_dir + "\\" + output_dir # Output directory is relative to the project directory
      generateBomCsv (output_dir = output_dir, sch_filename = sch_file_path)

    elif cmd == "pcb_pdf":
      output_dir = current_config.get ("data", {}).get ("pcb_pdf", {}).get ("--output_dir", lambda: default_config ["data"]["pcb_pdf"]["--output_dir"])
      output_dir = project_dir + "\\" + output_dir # Output directory is relative to the project directory
      generatePcbPdf (output_dir = output_dir, pcb_filename = pcb_file_path)

    elif cmd == "pcb_fab_pdf":
      output_dir = current_config.get ("data", {}).get ("pcb_fab_pdf", {}).get ("--output_dir", lambda: default_config ["data"]["pcb_fab_pdf"]["--output_dir"])
      output_dir = project_dir + "\\" + output_dir # Output directory is relative to the project directory
      generatePcbFabPdf (output_dir = output_dir, pcb_filename = pcb_file_path)

    elif cmd == "positions":
      output_dir = current_config.get ("data", {}).get ("positions", {}).get ("--output_dir", lambda: default_config ["data"]["positions"]["--output_dir"])
      output_dir = project_dir + "\\" + output_dir # Output directory is relative to the project directory
      generatePositions (output_dir = output_dir, pcb_filename = pcb_file_path)

    elif cmd == "ddd":
      # Default is STEP
      output_dir = current_config.get ("data", {}).get ("ddd", {}).get ("STEP", {}).get ("--output_dir", lambda: default_config ["data"]["ddd"]["STEP"]["--output_dir"])
      output_dir = project_dir + "\\" + output_dir # Output directory is relative to the project directory
      generate3D (output_dir = output_dir, pcb_filename = pcb_file_path, type = "STEP")
    
    elif cmd == "svg":
      output_dir = current_config.get ("data", {}).get ("svg", {}).get ("--output_dir", lambda: default_config ["data"]["svg"]["--output_dir"])
      output_dir = project_dir + "\\" + output_dir # Output directory is relative to the project directory
      generateSvg (output_dir = output_dir, pcb_filename = pcb_file_path)

    elif cmd == "pcb_render":
      output_dir = current_config.get ("data", {}).get ("pcb_render", {}).get ("--output_dir", lambda: default_config ["data"]["pcb_render"]["--output_dir"])
      output_dir = project_dir + "\\" + output_dir # Output directory is relative to the project directory
      generatePcbRenders (output_dir = output_dir, pcb_filename = pcb_file_path)

    elif cmd == "source_zip":
      src_output_dir = current_config.get ("data", {}).get ("source_zip", {}).get ("--output_dir", lambda: default_config ["data"]["source_zip"]["--output_dir"])
      src_output_dir = project_dir + "\\" + src_output_dir # Output directory is relative to the project directory

  #---------------------------------------------------------------------------------------------#

  # Process the commands with arguments or modifiers. eg. "["ddd", "STEP"]", "[ddd, VRML]"
  for cmd in cmd_lists:
    if cmd [0] == "pcb_drc":
      if cmd [1] == "report":
        output_dir = current_config.get ("data", {}).get ("pcb_drc", {}).get ("--output_dir", lambda: default_config ["data"]["pcb_drc"]["--output_dir"])
        output_dir = project_dir + "\\" + output_dir # Output directory is relative to the project directory
        runDRC (output_dir = output_dir, pcb_filename = pcb_file_path, type = "report")
      
      elif cmd [1] == "json":
        output_dir = current_config.get ("data", {}).get ("pcb_drc", {}).get ("--output_dir", lambda: default_config ["data"]["pcb_drc"]["--output_dir"])
        output_dir = project_dir + "\\" + output_dir # Output directory is relative to the project directory
        runDRC (output_dir = output_dir, pcb_filename = pcb_file_path, type = "json")

    elif cmd [0] == "gerbers":
      output_dir = current_config.get ("data", {}).get ("gerbers", {}).get ("--output_dir", lambda: default_config ["data"]["gerbers"]["--output_dir"])
      output_dir = project_dir + "\\" + output_dir # Output directory is relative to the project directory
      generateGerbers (output_dir = output_dir, pcb_filename = pcb_file_path)
   
    elif cmd [0] == "drills":
      output_dir = current_config.get ("data", {}).get ("drills", {}).get ("--output_dir", lambda: default_config ["data"]["drills"]["--output_dir"])
      output_dir = project_dir + "\\" + output_dir # Output directory is relative to the project directory
      generateDrills (output_dir = output_dir, pcb_filename = pcb_file_path)

    elif cmd [0] == "sch_erc":
      if cmd [1] == "report":
        output_dir = current_config.get ("data", {}).get ("sch_erc", {}).get ("--output_dir", lambda: default_config ["data"]["sch_erc"]["--output_dir"])
        output_dir = project_dir + "\\" + output_dir # Output directory is relative to the project directory
        runERC (output_dir = output_dir, sch_filename = sch_file_path, type = "report")
      
      elif cmd [1] == "json":
        output_dir = current_config.get ("data", {}).get ("sch_erc", {}).get ("--output_dir", lambda: default_config ["data"]["sch_erc"]["--output_dir"])
        output_dir = project_dir + "\\" + output_dir # Output directory is relative to the project directory
        runERC (output_dir = output_dir, sch_filename = sch_file_path, type = "json")

    elif cmd [0] == "sch_pdf":
      output_dir = current_config.get ("data", {}).get ("sch_pdf", {}).get ("--output_dir", lambda: default_config ["data"]["sch_pdf"]["--output_dir"])
      output_dir = project_dir + "\\" + output_dir # Output directory is relative to the project directory
      generateSchPdf (output_dir = output_dir, sch_filename = sch_file_path)
    
    elif cmd [0] == "bom":
      if cmd [1] == "CSV":
        if csv_file_created == False:
          # Xls below may have generated a CSV first - generate CSV ONLY if not already generated by below XLS
          output_dir = current_config.get ("data", {}).get ("bom", {}).get ("CSV", {}).get ("--output_dir", lambda: default_config ["data"]["bom"]["CSV"]["--output_dir"])
          output_dir = project_dir + "\\" + output_dir # Output directory is relative to the project directory
          csv_filename = generateBomCsv (output_dir = output_dir, sch_filename = sch_file_path)
          csv_file_created = True

      if cmd [1] == "XLS":
        # Xls generation need to have a CSV first - generate CSV ONLY if not already generated by user request
        if csv_file_created == False:
          output_dir = current_config.get ("data", {}).get ("bom", {}).get ("CSV", {}).get ("--output_dir", lambda: default_config ["data"]["bom"]["CSV"]["--output_dir"])
          output_dir = project_dir + "\\" + output_dir # Output directory is relative to the project directory
          csv_filename = generateBomCsv (output_dir = output_dir, sch_filename = sch_file_path)
          csv_file_created = True
        output_dir = current_config.get ("data", {}).get ("bom", {}).get ("XLS", {}).get ("--output_dir", lambda: default_config ["data"]["bom"]["XLS"]["--output_dir"])
        output_dir = project_dir + "\\" + output_dir # Output directory is relative to the project directory
        generateBomXls (output_dir = output_dir, csv_file = csv_filename, sch_filename = sch_file_path)
    
      elif cmd [1] == "HTML":
        output_dir = current_config.get ("data", {}).get ("bom", {}).get ("HTML", {}).get ("--output_dir", lambda: default_config ["data"]["bom"]["HTML"]["--output_dir"])
        output_dir = project_dir + "\\" + output_dir # Output directory is relative to the project directory
        generateBomHtml (output_dir = output_dir, pcb_filename = pcb_file_path)
    
    elif cmd [0] == "pcb_pdf":
      output_dir = current_config.get ("data", {}).get ("pcb_pdf", {}).get ("--output_dir", lambda: default_config ["data"]["pcb_pdf"]["--output_dir"])
      output_dir = project_dir + "\\" + output_dir # Output directory is relative to the project directory
      generatePcbPdf (output_dir = output_dir, pcb_filename = pcb_file_path)
    
    elif cmd [0] == "pcb_fab_pdf":
      output_dir = current_config.get ("data", {}).get ("pcb_fab_pdf", {}).get ("--output_dir", lambda: default_config ["data"]["pcb_fab_pdf"]["--output_dir"])
      output_dir = project_dir + "\\" + output_dir # Output directory is relative to the project directory
      generatePcbFabPdf (output_dir = output_dir, pcb_filename = pcb_file_path)
    
    elif cmd [0] == "positions":
      output_dir = current_config.get ("data", {}).get ("positions", {}).get ("--output_dir", lambda: default_config ["data"]["positions"]["--output_dir"])
      output_dir = project_dir + "\\" + output_dir # Output directory is relative to the project directory
      generatePositions (output_dir = output_dir, pcb_filename = pcb_file_path)
    
    elif cmd [0] == "ddd":
      if cmd [1] == "VRML":
        output_dir = current_config.get ("data", {}).get ("ddd", {}).get ("VRML", {}).get ("--output_dir", lambda: default_config ["data"]["ddd"]["VRML"]["--output_dir"])
        output_dir = project_dir + "\\" + output_dir # Output directory is relative to the project directory
        generate3D (output_dir = output_dir, pcb_filename = pcb_file_path, type = "VRML")
        
      else: # Default is STEP
        output_dir = current_config.get ("data", {}).get ("ddd", {}).get ("STEP", {}).get ("--output_dir", lambda: default_config ["data"]["ddd"]["STEP"]["--output_dir"])
        output_dir = project_dir + "\\" + output_dir # Output directory is relative to the project directory
        generate3D (output_dir = output_dir, pcb_filename = pcb_file_path, type = "STEP")
    
    elif cmd [0] == "svg":
      output_dir = current_config.get ("data", {}).get ("svg", {}).get ("--output_dir", lambda: default_config ["data"]["svg"]["--output_dir"])
      output_dir = project_dir + "\\" + output_dir # Output directory is relative to the project directory
      generateSvg (output_dir = output_dir, pcb_filename = pcb_file_path)
    
    elif cmd [0] == "pcb_render":
      output_dir = current_config.get ("data", {}).get ("pcb_render", {}).get ("--output_dir", lambda: default_config ["data"]["pcb_render"]["--output_dir"])
      output_dir = project_dir + "\\" + output_dir # Output directory is relative to the project directory
      # Run for the items in the list by iterating from the second item.
      for preset in cmd [1:]:
        generatePcbRenders (output_dir = output_dir, pcb_filename = pcb_file_path, preset = preset)
  
    elif cmd [0] == "source_zip":
      src_output_dir = current_config.get ("data", {}).get ("source_zip", {}).get ("--output_dir", lambda: default_config ["data"]["source_zip"]["--output_dir"])
      src_output_dir = project_dir + "\\" + src_output_dir # Output directory is relative to the project directory
    
  # Once the command execution is complete, print the statuses from command_exec_status.
  print (f"run [INFO]: Command execution completed. Statuses:")
  command_exec_status_all = True
  for cmd in command_exec_status:
    if (command_exec_status [cmd] == True):
      print (color.green (f"run [INFO]: {cmd}: {command_exec_status [cmd]}"))
    else:
      print (color.red (f"run [INFO]: {cmd}: {command_exec_status [cmd]}"))
      command_exec_status_all = False

  # If requested source backup after all generation successful
  if src_output_dir != None:
    if command_exec_status_all == True:
      generateSourceZip (output_dir = src_output_dir, pcb_filename = pcb_file_path, sch_filename = sch_file_path)
    else:
      print (color.yellow ("run [WARNING]: Source file zip not created because some commands generation failed."))
   
  return

#=============================================================================================#

def generateSourceZip (output_dir, pcb_filename = None, sch_filename = None):
  """
  Generate the SRC directory with zip of all relevnt source files.
  """
  global current_config  # Access the global config
  global default_config  # Access the global config

  # Prefer using pcb_filename if provided.
  if pcb_filename == None:
    file_name = sch_filename
  else:
    file_name = pcb_filename
  file_path = os.path.abspath (file_name) # Get the absolute path of the file.
  # Check if the input file exists
  if not check_file_exists (file_name):
    print (color.red (f"generateSourceZip [ERROR]: '{file_name}' does not exist."))
    return
  
  #---------------------------------------------------------------------------------------------#

  file_name = extract_pcb_file_name (file_name) # Extract information from the input file
  file_name = file_name.replace (" ", "-") # If there are whitespace characters in the project name, replace them with a hyphen

  project_name = extract_project_name (file_name)
  info = extract_info_from_pcb (pcb_filename) # Extract basic information from the input file

  print (f"generateSourceZip [INFO]: Project name is '{color.magenta (project_name)}' and revision is {color.magenta ('R')}{color.magenta (info ['rev'])}.")

  #---------------------------------------------------------------------------------------------#

  # Get the directory path of the file and save it as the project directory.
  # All other export directories will be relative to the project directory.
  project_dir = os.path.dirname (file_path)
  
  # Read the output directory name from the config file.
  od_from_config = project_dir + "/" + current_config.get ("data", {}).get ("source_zip", {}).get ("--output_dir", lambda: default_config ["data"]["source_zip"]["--output_dir"])
  od_from_cli = output_dir  # The output directory specified by the command line argument

  # Get the final directory path.
  final_directory, filename_date = create_final_directory (od_from_config, od_from_cli, "SRC", info ["rev"], "generateSourceZip")
  
  #---------------------------------------------------------------------------------------------#

  # list of project source files to include in zip file 
  ext_list = current_config.get ("data", {}).get ("source_zip", {}).get ("kie_ext_list", lambda: default_config ["data"]["source_zip"]["kie_ext_list"])
  files_list = []
  for ext in ext_list:
    if ext[0] == '.':
      files_list.append( f'{project_name}{ext}' )
    else:
      files_list.append( f'{project_name}.{ext}' )

  not_completed = True
  # Sequentially name and create the zip files.
  seq_number = 1
  while not_completed:
    zip_file_name = f"{project_name}-R{info ['rev']}-Source-{filename_date}-{seq_number}.zip"
    zip_file_path = f"{final_directory}/{zip_file_name}"
    if os.path.exists (zip_file_path):
      seq_number += 1
    else:
      zip_cnt = zip_all_files_list (project_dir, zip_file_path, files_list)
      print (f"generateSourceZip [OK]: ZIP file '{color.magenta (zip_file_name)}' created successfully with {zip_cnt} Files.")
      print()
      not_completed = False

#=============================================================================================#

def runDRC (output_dir, pcb_filename, type = "default"):
  """
  Runs the DRC (Design Rule Check) on the PCB file.
  This function is a placeholder and should be implemented with actual DRC logic.
  """
  global current_config  # Access the global config
  global default_config  # Access the global config

  # Get the KiCad CLI path.
  kicad_cli_path = f'{current_config.get ("kicad_cli_path", lambda: default_config ["kicad_cli_path"])}'

  # Common base command
  pcb_drc_command = [f'"{kicad_cli_path}"', "pcb", "drc"]

  # Check if the input file exists
  if not check_file_exists (pcb_filename):
    print (color.red (f"runDRC [ERROR]: '{pcb_filename}' does not exist."))
    command_exec_status ["pcb_drc"] = False
    return
  
  #---------------------------------------------------------------------------------------------#

  file_name = extract_pcb_file_name (pcb_filename) # Extract information from the input file
  file_name = file_name.replace (" ", "-") # If there are whitespace characters in the project name, replace them with a hyphen

  project_name = extract_project_name (file_name)
  info = extract_info_from_pcb (pcb_filename) # Extract basic information from the input file

  print (f"runDRC [INFO]: Project name is '{color.magenta (project_name)}' and revision is {color.magenta ('R')}{color.magenta (info ['rev'])}.")

  #---------------------------------------------------------------------------------------------#

  # print (color.yellow ("runDRC [INFO]: Running DRC (Design Rule Check) on the PCB file.."))

  file_path = os.path.abspath (pcb_filename) # Get the absolute path of the file.

  # Get the directory path of the file and save it as the project directory.
  # All other export directories will be relative to the project directory.
  project_dir = os.path.dirname (file_path)
  
  # Read the output directory name from the config file.
  od_from_config = project_dir + "/" + current_config.get ("data", {}).get ("pcb_drc", {}).get ("--output_dir", lambda: default_config ["data"]["pcb_drc"]["--output_dir"])
  od_from_cli = output_dir  # The output directory specified by the command line argument

  # Get the final directory path.
  final_directory, filename_date = create_final_directory (od_from_config, od_from_cli, "Report", info ["rev"], "runDRC")
  
  #---------------------------------------------------------------------------------------------#

  # Get the argument list from the config file.
  arg_list = current_config.get ("data", {}).get ("pcb_drc", {})

  full_command = [] # Store the full command to run
  full_command.extend (pcb_drc_command) # Add the base command

  if type == "default": # If the default type is specified, then use type specified in the config file.
    if arg_list: # Check if the argument list is not empty.
      if "--format" in arg_list: # Check if the "--format" argument is present in the argument list.
        format_value = arg_list ["--format"] # Get the format value from the argument list.
        if format_value == "report":
          type = format_value
        elif format_value == "json":
          type = format_value
      else:
        type = "report"  # Default to report if no format is specified in the config file.
    else: # If the argument list is empty, then default to report.
      type = "report"  # Default to report if no argument list is provided.
  else:
    pass

  #---------------------------------------------------------------------------------------------#

  seq_number = 1
  not_completed = True
  
  # Create the output file name.
  while not_completed:
    if type == "report":
      file_name = f"{final_directory}/{project_name}-R{info ['rev']}-PCB-DRC-Report-{filename_date}-{seq_number}.rpt"
    elif type == "json":
      file_name = f"{final_directory}/{project_name}-R{info ['rev']}-PCB-DRC-Report-{filename_date}-{seq_number}.json"

    if os.path.exists (file_name):
      seq_number += 1 # Increment the sequence number and try again
      not_completed = True
    else:
      full_command.append ("--output")
      full_command.append (f'"{file_name}"') # Add the output file name with double quotes around it
      break
  
  # Add the remaining arguments.
  # Check if the argument list is not an empty dictionary.
  if arg_list:
    for key, value in arg_list.items():
      if key.startswith ("--"): # Only fetch the arguments that start with "--"
        if key == "--output_dir": # Skip the --output_dir argument, sice we already added it
          continue
        elif key == "--format": # Skip the --format argument, since we already set the type
          if isinstance (value, str):
            full_command.append (key)
            full_command.append (f'"{type}"') # Use the override value. Add as a double-quoted string.
        else:
          # Check if the value is empty
          if value == "": # Skip if the value is empty
            continue
          else:
            # Check if the vlaue is a JSON boolean
            if isinstance (value, bool):
              if value == True: # If the value is true, then append the key as an argument
                full_command.append (key)
            else:
              # Check if the value is a string and not a numeral
              if isinstance (value, str) and not value.isdigit():
                  full_command.append (key)
                  full_command.append (f'"{value}"') # Add as a double-quoted string
              elif isinstance (value, (int, float)):
                  full_command.append (key)
                  full_command.append (str (value))  # Append the numeric value as string
  
  # Finally add the input file
  full_command.append (f'"{pcb_filename}"')
  print ("runDRC [INFO]: Running command: ", color.blue (' '.join (full_command)))

  #----------------------------------------------------------------------------------------------#

  # Run the command
  try:
    full_command = ' '.join (full_command) # Convert the list to a string
    result = subprocess.run (full_command, check = True, capture_output = True, text = True)

    print (result.stdout)  # Print the standard output

    is_drc_error = False

    # Check for a 0 DRC violations.
    if "Found 0 violations" in result.stdout:
      print (color.green ("runDRC [OK]: No DRC violations found."))
    
    # Search for the "Found <number> violations" string in the output using regex.
    elif re.search (r"Found \d+ violations", result.stdout):
      violations = re.search (r"Found (\d+) violations", result.stdout)
      if violations:
        num_violations = violations.group (1)
        print (color.red (f"runDRC [ERROR]: Found {num_violations} DRC violations. Check the report file for details."))
        is_drc_error = True
    
    if "Found 0 unconnected items" in result.stdout:
      print (color.green ("runDRC [OK]: No unconnected items found."))

    # Search for the "Found <number> unconnected items" string in the output using regex.
    elif re.search (r"Found \d+ unconnected items", result.stdout):
      unconnected = re.search (r"Found (\d+) unconnected items", result.stdout)
      if unconnected:
        num_unconnected = unconnected.group (1)
        print (color.red (f"runDRC [ERROR]: Found {num_unconnected} unconnected items. Check the report file for details."))
        is_drc_error = True
    
    if "Found 0 schematic parity issues" in result.stdout:
      print (color.green ("runDRC [OK]: No schematic parity issues found."))
    
    # Search for the "Found <number> schematic parity issues" string in the output using regex.
    elif re.search (r"Found \d+ schematic parity issues", result.stdout):
      parity_issues = re.search (r"Found (\d+) schematic parity issues", result.stdout)
      if parity_issues:
        num_parity_issues = parity_issues.group (1)
        print (color.red (f"runDRC [ERROR]: Found {num_parity_issues} schematic parity issues. Check the report file for details."))
        is_drc_error = True
  
    if is_drc_error:
      print (color.yellow ("runDRC [WARNING]: Multiple DRC errors were found. Do you want to continue? [Y/N]"))
      user_input = input ("").strip().lower()
      if user_input == "y":
        print (color.yellow ("runDRC [WARNING]: Process will continue with DRC errors."))
        command_exec_status ["pcb_drc"] = False
      else:
        print (color.yellow ("runDRC [WARNING]: Process will exit."))
        exit (1)  # Exit the script with an error code if the user does not want to continue

  except subprocess.CalledProcessError as e: # If the DRC command fails for some reason.
    print (color.red (f"runDRC [ERROR]: Error occurred: {e}"))
    print()
    command_exec_status ["pcb_drc"] = False
    return

  print (color.green ("runDRC [OK]: DRC completed successfully."))
  print()
  command_exec_status ["pcb_drc"] = True

#=============================================================================================#

def runERC (output_dir, sch_filename, type = "default"):
  """
  Runs the ERC (Electrical Rule Check) on the SCH file.
  This function is a placeholder and should be implemented with actual ERC logic.
  """
  global current_config  # Access the global config
  global default_config  # Access the global config

  # Get the KiCad CLI path.
  kicad_cli_path = f'{current_config.get ("kicad_cli_path", lambda: default_config ["kicad_cli_path"])}'

  # Common base command
  sch_erc_command = [f'"{kicad_cli_path}"', "sch", "erc"]

  # Check if the input file exists
  if not check_file_exists (sch_filename):
    print (color.red (f"runERC [ERROR]: '{sch_filename}' does not exist."))
    command_exec_status ["sch_erc"] = False
    return
  
  #---------------------------------------------------------------------------------------------#

  file_name = extract_pcb_file_name (sch_filename) # Extract information from the input file
  file_name = file_name.replace (" ", "-") # If there are whitespace characters in the project name, replace them with a hyphen

  project_name = extract_project_name (file_name)
  info = extract_info_from_pcb (sch_filename) # Extract basic information from the input file

  print (f"runERC [INFO]: Project name is '{color.magenta (project_name)}' and revision is {color.magenta ('R')}{color.magenta (info ['rev'])}.")

  #---------------------------------------------------------------------------------------------#

  # print (color.yellow ("runERC [INFO]: Running ERC (Electrical Rule Check) on the SCH file.."))

  file_path = os.path.abspath (sch_filename) # Get the absolute path of the file.

  # Get the directory path of the file and save it as the project directory.
  # All other export directories will be relative to the project directory.
  project_dir = os.path.dirname (file_path)
  
  # Read the output directory name from the config file.
  od_from_config = project_dir + "/" + current_config.get ("data", {}).get ("sch_erc", {}).get ("--output_dir", lambda: default_config ["data"]["sch_erc"]["--output_dir"])
  od_from_cli = output_dir  # The output directory specified by the command line argument

  # Get the final directory path.
  final_directory, filename_date = create_final_directory (od_from_config, od_from_cli, "Report", info ["rev"], "runERC")
  
  #---------------------------------------------------------------------------------------------#

  # Get the argument list from the config file.
  arg_list = current_config.get ("data", {}).get ("sch_erc", {})

  full_command = [] # Store the full command to run
  full_command.extend (sch_erc_command) # Add the base command

  if type == "default": # If the default type is specified, then use type specified in the config file.
    if arg_list: # Check if the argument list is not empty.
      if "--format" in arg_list: # Check if the "--format" argument is present in the argument list.
        format_value = arg_list ["--format"] # Get the format value from the argument list.
        if format_value == "report":
          type = format_value
        elif format_value == "json":
          type = format_value
      else:
        type = "report"  # Default to report if no format is specified in the config file.
    else: # If the argument list is empty, then default to report.
      type = "report"  # Default to report if no argument list is provided.
  else:
    pass

  #---------------------------------------------------------------------------------------------#

  seq_number = 1
  not_completed = True
  
  # Create the output file name.
  while not_completed:
    if type == "report":
      file_name = f"{final_directory}/{project_name}-R{info ['rev']}-SCH-ERC-Report-{filename_date}-{seq_number}.rpt"
    elif type == "json":
      file_name = f"{final_directory}/{project_name}-R{info ['rev']}-SCH-ERC-Report-{filename_date}-{seq_number}.json"

    if os.path.exists (file_name):
      seq_number += 1 # Increment the sequence number and try again
      not_completed = True
    else:
      full_command.append ("--output")
      full_command.append (f'"{file_name}"') # Add the output file name with double quotes around it
      break
  
  # Add the remaining arguments.
  # Check if the argument list is not an empty dictionary.
  if arg_list:
    for key, value in arg_list.items():
      if key.startswith ("--"): # Only fetch the arguments that start with "--"
        if key == "--output_dir": # Skip the --output_dir argument, sice we already added it
          continue
        elif key == "--format": # Skip the --format argument, since we already set the type
          if isinstance (value, str):
            full_command.append (key)
            full_command.append (f'"{type}"') # Use the override value. Add as a double-quoted string.
        else:
          # Check if the value is empty
          if value == "": # Skip if the value is empty
            continue
          else:
            # Check if the vlaue is a JSON boolean
            if isinstance (value, bool):
              if value == True: # If the value is true, then append the key as an argument
                full_command.append (key)
            else:
              # Check if the value is a string and not a numeral
              if isinstance (value, str) and not value.isdigit():
                  full_command.append (key)
                  full_command.append (f'"{value}"') # Add as a double-quoted string
              elif isinstance (value, (int, float)):
                  full_command.append (key)
                  full_command.append (str (value))  # Append the numeric value as string
  
  # Finally add the input file
  full_command.append (f'"{sch_filename}"')
  print ("runERC [INFO]: Running command: ", color.blue (' '.join (full_command)))

  #----------------------------------------------------------------------------------------------#

  # Run the command
  try:
    full_command = ' '.join (full_command) # Convert the list to a string
    result = subprocess.run (full_command, check = True, capture_output = True, text = True)

    print (result.stdout)  # Print the standard output

    is_erc_error = False

    # Check for a 0 ERC violations.
    if "Found 0 violations" in result.stdout:
      print (color.green ("runERC [OK]: No ERC violations found."))
    
    # Search for the "Found <number> violations" string in the output using regex.
    elif re.search (r"Found \d+ violations", result.stdout):
      violations = re.search (r"Found (\d+) violations", result.stdout)
      if violations:
        num_violations = violations.group (1)
        print (color.red (f"runERC [ERROR]: Found {num_violations} ERC violations. Check the report file for details."))
        is_erc_error = True
    
    if is_erc_error:
      print (color.yellow ("runERC [WARNING]: Multiple ERC errors were found. Do you want to continue? [Y/N]"))
      user_input = input ("").strip().lower()
      if user_input == "y":
        print (color.yellow ("runERC [WARNING]: Process will continue with ERC errors."))
        command_exec_status ["pcb_erc"] = False
      else:
        print (color.yellow ("runERC [WARNING]: Process will exit."))
        exit (1)  # Exit the script with an error code if the user does not want to continue

  except subprocess.CalledProcessError as e: # If the ERC command fails for some reason.
    print (color.red (f"runERC [ERROR]: Error occurred: {e}"))
    print()
    command_exec_status ["sch_erc"] = False
    return

  print (color.green ("runERC [OK]: ERC completed successfully."))
  print()
  command_exec_status ["sch_erc"] = True

#=============================================================================================#

def test():
  info = extract_info_from_pcb (SAMPLE_PCB_FILE)
  print (info)
  print (f"Revision is {info ['rev']}")

#=============================================================================================#

def parseArguments():
  """
  Parses commands and arguments, and execute the corresponding functions.
  """
  # Configure the argument parser.
  parser = argparse.ArgumentParser (description = f"{APP_NAME}: {APP_DESCRIPTION}")
  parser.add_argument ('-v', '--version', action = 'version', version = f'{APP_VERSION}', help = "Show the version of the tool and exit.")
  subparsers = parser.add_subparsers (dest = "command", help = "Available commands.")

  # Subparser for the Run command.
  # Example: python .\kiexport.py run "Mitayi-Pico-D1/kiexport.json"
  run_parser = subparsers.add_parser ("run", help = "Run KiExport using the provided JSON configuration file.")
  run_parser.add_argument ("config_file", help = "Path to the JSON configuration file.")
  run_parser.add_argument ("command_list", nargs = "?", help = "Specific commands in the JSON to execute (optional).")
  run_parser.add_argument ("-c", "--clean", action="store_true", help = "Force clean version directory.")

  # Subparser for the Gerber export command.
  # Example: python .\kiexport.py gerbers -od "Mitayi-Pico-D1/Export" -if "Mitayi-Pico-D1/Mitayi-Pico-RP2040.kicad_pcb"
  gerbers_parser = subparsers.add_parser ("gerbers", help = "Export Gerber files.")
  gerbers_parser.add_argument ("-if", "--input_filename", required = True, help = "Path to the .kicad_pcb file.")
  gerbers_parser.add_argument ("-od", "--output_dir", required = True, help = "Directory to save the Gerber files to.")

  # Subparser for the Drills export command.
  # Example: python .\kiexport.py drills -od "Mitayi-Pico-D1/Export" -if "Mitayi-Pico-D1/Mitayi-Pico-RP2040.kicad_pcb"
  drills_parser = subparsers.add_parser ("drills", help = "Export Drill files.")
  drills_parser.add_argument ("-if", "--input_filename", required = True, help = "Path to the .kicad_pcb file.")
  drills_parser.add_argument ("-od", "--output_dir", required = True, help = "Directory to save the Drill files to.")

  # Subparser for the Position file export command.
  # Example: python .\kiexport.py positions -od "Mitayi-Pico-D1/Export" -if "Mitayi-Pico-D1/Mitayi-Pico-RP2040.kicad_pcb"
  positions_parser = subparsers.add_parser ("positions", help = "Export Position files.")
  positions_parser.add_argument ("-if", "--input_filename", required = True, help = "Path to the .kicad_pcb file.")
  positions_parser.add_argument ("-od", "--output_dir", required = True, help = "Directory to save the Position files to.")

  # Subparser for the PCB PDF export command.
  # Example: python .\kiexport.py pcb_pdf -od "Mitayi-Pico-D1/Export" -if "Mitayi-Pico-D1/Mitayi-Pico-RP2040.kicad_pcb"
  pcb_pdf_parser = subparsers.add_parser ("pcb_pdf", help = "Export PCB PDF files.")
  pcb_pdf_parser.add_argument ("-if", "--input_filename", required = True, help = "Path to the .kicad_pcb file.")
  pcb_pdf_parser.add_argument ("-od", "--output_dir", required = True, help = "Directory to save the PCB PDF files to.")

  # Subparser for the PCB Fab PDF export command.
  # Example: python .\kiexport.py pcb_pdf -od "Mitayi-Pico-D1/Export" -if "Mitayi-Pico-D1/Mitayi-Pico-RP2040.kicad_pcb"
  pcb_pdf_parser = subparsers.add_parser ("pcb_fab_pdf", help = "Export PCB Fab PDF files.")
  pcb_pdf_parser.add_argument ("-if", "--input_filename", required = True, help = "Path to the .kicad_pcb file.")
  pcb_pdf_parser.add_argument ("-od", "--output_dir", required = True, help = "Directory to save the PCB Fab PDF files to.")

  # Subparser for the Schematic PDF export command.
  # Example: python .\kiexport.py sch_pdf -od "Mitayi-Pico-D1/Export" -if "Mitayi-Pico-D1/Mitayi-Pico-RP2040.kicad_sch"
  sch_pdf_parser = subparsers.add_parser ("sch_pdf", help = "Export schematic PDF files.")
  sch_pdf_parser.add_argument ("-if", "--input_filename", required = True, help = "Path to the .kicad_sch file.")
  sch_pdf_parser.add_argument ("-od", "--output_dir", required = True, help = "Directory to save the Schematic PDF files to.")

  # Subparser for the 3D file export command.
  # Example: python .\kiexport.py ddd -t "VRML" -od "Mitayi-Pico-D1/Export" -if "Mitayi-Pico-D1/Mitayi-Pico-RP2040.kicad_pcb"
  ddd_parser = subparsers.add_parser ("ddd", help = "Export 3D files.")
  ddd_parser.add_argument ("-if", "--input_filename", required = True, help = "Path to the .kicad_pcb file.")
  ddd_parser.add_argument ("-od", "--output_dir", required = True, help = "Directory to save the 3D files to.")
  ddd_parser.add_argument ("-t", "--type", required = True, help = "The type of file to generate. Can be STEP or VRML.")

  # Subparser for the BoM file export command.
  # Example: python .\kiexport.py bom -od "Mitayi-Pico-D1/Export" -if "Mitayi-Pico-D1/Mitayi-Pico-RP2040.kicad_sch" -t "CSV"
  bom_parser = subparsers.add_parser ("bom", help = "Export BoM files.")
  bom_parser.add_argument ("-if", "--input_filename", required = True, help = "Path to the .kicad_sch file.")
  bom_parser.add_argument ("-od", "--output_dir", required = True, help = "Directory to save the BoM files to.")
  bom_parser.add_argument ("-t", "--type", help = "The type of file to generate. Default is CSV.")

  # # Subparser for the HTML iBoM file export command.
  # # Example: python .\kiexport.py ibom -od "Mitayi-Pico-D1/Export" -if "Mitayi-Pico-D1/Mitayi-Pico-RP2040.kicad_pcb"
  # ibom_parser = subparsers.add_parser ("ibom", help = "Export HMTL iBoM files. The Kicad iBOM plugin is required")
  # ibom_parser.add_argument ("-if", "--input_filename", required = True, help = "Path to the .kicad_pcb file.")
  # ibom_parser.add_argument ("-od", "--output_dir", required = True, help = "Directory to save the BoM files to.")

  # Subparser for the SVG export command.
  # Example: python .\kiexport.py svg -od "Mitayi-Pico-D1/Export" -if "Mitayi-Pico-D1/Mitayi-Pico-RP2040.kicad_pcb"
  svg_parser = subparsers.add_parser ("svg", help = "Export SVG files.")
  svg_parser.add_argument ("-if", "--input_filename", required = True, help = "Path to the .kicad_pcb file.")
  svg_parser.add_argument ("-od", "--output_dir", required = True, help = "Directory to save the SVG files to.")

  # Subparser for the PCB Render export command.
  # Example: python .\kiexport.py pcb_render -od "Mitayi-Pico-D1/Export" -if "Mitayi-Pico-D1/Mitayi-Pico-RP2040.kicad_pcb"
  pcb_render_parser = subparsers.add_parser ("pcb_render", help = "Export rendered PCB images.")
  pcb_render_parser.add_argument ("-if", "--input_filename", required = True, help = "Path to the .kicad_pcb file.")
  pcb_render_parser.add_argument ("-od", "--output_dir", required = True, help = "Directory to save the rendered images to.")
  pcb_render_parser.add_argument ("-ps", "--preset", required = False, help = "The render preset to use.")

  # Subparser for the DRC Run command.
  # Example: python .\kiexport.py pcb_drc -od "Mitayi-Pico-D1/Export" -if "Mitayi-Pico-D1/Mitayi-Pico-RP2040.kicad_pcb" -t "report"
  pcb_drc_parser = subparsers.add_parser ("pcb_drc", help = "Run DRC on the PCB file and generate a report.")
  pcb_drc_parser.add_argument ("-if", "--input_filename", required = True, help = "Path to the .kicad_pcb file.")
  pcb_drc_parser.add_argument ("-od", "--output_dir", required = True, help = "Directory to save the report file to.")
  pcb_drc_parser.add_argument ("-t", "--type", required = False, help = "The type of report file. Can be report or json.")

  # Subparser for the ERC Run command.
  # Example: python .\kiexport.py sch_erc -od "Mitayi-Pico-D1/Export" -if "Mitayi-Pico-D1/Mitayi-Pico-RP2040.kicad_pcb" -t "report"
  sch_erc_parser = subparsers.add_parser ("sch_erc", help = "Run ERC on the SCH file and generate a report.")
  sch_erc_parser.add_argument ("-if", "--input_filename", required = True, help = "Path to the .kicad_pcb file.")
  sch_erc_parser.add_argument ("-od", "--output_dir", required = True, help = "Directory to save the report file to.")
  sch_erc_parser.add_argument ("-t", "--type", required = False, help = "The type of report file. Can be report or json.")

  # Subparser for the test function.
  test_parser = subparsers.add_parser ("test", help = "Internal test function.")

  #---------------------------------------------------------------------------------------------#

  # Parse arguments.
  args = parser.parse_args()
  printInfo()

  #---------------------------------------------------------------------------------------------#

  # Handle empty arguments.
  if args.command is None:
    print (color.red ("Looks like you forgot to specify any inputs. Time to RTFM."))
    print()
    parser.print_help()
    exit()
  
  #---------------------------------------------------------------------------------------------#

  # Handle the version command.
  if args.command == "-v" or args.command == "--version":
    print (f"KiExport v{APP_VERSION}")
    return

  #---------------------------------------------------------------------------------------------#

  # The Run command accepts a config file as an argument and generate the files based on the
  # config file. The name of the config file can be anything.
  if args.command == "run":
    # Check if the paramter ends with ".json"
    if args.config_file.endswith (".json") or args.config_file.endswith (".json\""):
      run (config_file = args.config_file, command_list = args.command_list, clean = args.clean)

  else:
    # Load the standard config file for other commands.
    if args.input_filename is not None: # Check if we received an input file
      config_file_path = os.path.join (os.path.dirname (args.input_filename), "kiexport.json")
      load_config (config_file = config_file_path)
    else:
      load_config (config_file = "kiexport.json")
  
  #---------------------------------------------------------------------------------------------#
  # Create the path for the log file.

  global logPath

  # Check if the loaded configuration has a log file path.
  path_from_config = current_config.get ("kiexport_log_path", lambda: default_config ["kiexport_log_path"])

  if path_from_config is not None:
    if path_from_config.endswith (".log") or path_from_config.endswith (".txt"): # Check if the path ends with a file extension.
      pass
    else:
      path_from_config = path_from_config + ".log"

    if os.path.isabs (path_from_config): # Check if the path is an absolute path.
      logPath = path_from_config
    else: # If it is not absolute, we will add the project directory to it.
      if args.command == "run":
        file_path = os.path.abspath (args.config_file) # Get the absolute path of the config file.
      else:
        file_path = os.path.abspath (args.input_filename) # Get the absolute path of the input file.
      # Get the directory path of the file and save it as the project directory.
      # All other export directories will be relative to the project directory.
      project_dir = os.path.dirname (file_path)
      logPath = project_dir + "\\" + path_from_config

  # if args.output_dir is not None:
  #   logPath = os.path.join (os.path.dirname (args.input_filename), "kiexport.log")

  #---------------------------------------------------------------------------------------------#
  
  # Check the command and run it.
  if args.command == "run":
    pass

  elif args.command == "pcb_drc":
    if args.type is None:
      args.type = "default"  # Set the default type. This will use the type specified in the config file.
    runDRC (output_dir = args.output_dir, pcb_filename = args.input_filename, type = args.type)
  
  elif args.command == "gerbers":
    generateGerbers (output_dir = args.output_dir, pcb_filename = args.input_filename)

  elif args.command == "drills":
    generateDrills (output_dir = args.output_dir, pcb_filename = args.input_filename)
  
  elif args.command == "positions":
    generatePositions (output_dir = args.output_dir, pcb_filename = args.input_filename)
  
  elif args.command == "pcb_pdf":
    generatePcbPdf (output_dir = args.output_dir, pcb_filename = args.input_filename)

  elif args.command == "pcb_fab_pdf":
    generatePcbFabPdf (output_dir = args.output_dir, pcb_filename = args.input_filename)

  elif args.command == "sch_erc":
    if args.type is None:
      args.type = "default"  # Set the default type. This will use the type specified in the config file.
    runERC (output_dir = args.output_dir, sch_filename = args.input_filename, type = args.type)
  
  elif args.command == "sch_pdf":
    generateSchPdf (output_dir = args.output_dir, sch_filename = args.input_filename)
  
  #.............................................................................................#

  elif args.command == "bom":
    type = None
    if args.type is not None:
      # convert to string to uppercase
      type = args.type.upper()
    else:
      type = "CSV" # Default is CSV
    
    if type not in ["CSV", "XLS", "HTML"]:
      print (color.yellow (f"run [WARNING]: Invalid BoM type '{type}' specified. Defaulting to CSV."))
      type = "CSV"
    
    if (type == "CSV") or (type == "XLS"):
      # We need to generate the CSV file first.
      csv_file_name = generateBomCsv (output_dir = args.output_dir, sch_filename = args.input_filename)
      if type == "XLS":
        # The CSV file is used to generate the XLS file.
        generateBomXls (output_dir = args.output_dir, csv_file = csv_file_name, sch_filename = args.input_filename)
    elif type == "HTML":
      generateBomHtml (output_dir = args.output_dir, pcb_filename = args.input_filename)
  
  #.............................................................................................#

  elif args.command == "ddd":
    generate3D (output_dir = args.output_dir, pcb_filename = args.input_filename, type = args.type)

  elif args.command == "svg":
    generateSvg (output_dir = args.output_dir, pcb_filename = args.input_filename)

  elif args.command == "pcb_render":
    generatePcbRenders (output_dir = args.output_dir, pcb_filename = args.input_filename, preset = args.preset)

  elif args.command == "source_zip":
    generateSourceZip (output_dir = args.output_dir, pcb_filename = args.input_filename)
    
  elif args.command == "test":
    test()
    
  else:
    parser.print_help()

#=============================================================================================#

def printInfo():
  print ("")
  print (color.cyan (f"{APP_NAME} v{APP_VERSION}"))
  print (color.cyan ("CLI tool to export design and manufacturing files from KiCad projects."))
  print (color.cyan (f"Author: {APP_AUTHOR}"))
  print (color.cyan ("Contributors: Dominic Le Blanc (@domleblanc94)"))
  print ("")

#=============================================================================================#

def initLogger():
  global logger
  # sys.stdout = Logger ("Export/kiexport.log")
  logger = Logger ("Export/kiexport.log")
  sys.stdout = logger
  return logger

#=============================================================================================#

def main():
  global logger
  initLogger()
  parseArguments()
  logger.save_to_file (logPath)
  
#=============================================================================================#

if __name__ == "__main__":
  main()

#=============================================================================================#